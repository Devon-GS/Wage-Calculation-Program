import os
from datetime import datetime, time
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import sqlite3

# Loop through clock files and collect last 30 files
clock_list = []
for filename in os.listdir('../Uniclox/'):
    file = filename.replace(" ","")
    if 'TL' in file and file[-7:-4] != '000':
        clock_list.append(filename)
        
recent = clock_list[-100:]

# Loop though each clock file by line and append badge and times to dat_times list
date_times = []
for file in recent:    
    f = open('../Uniclox/' + file, 'r')
    for line in f:
        line = line.strip()
        line = line.split(',')
        badge = int(line[0])
        dated = line[1]
        timestamp = datetime.strptime(dated, '%Y-%m-%d %H:%M:%S').strftime("%d/%m/%y %H:%M:%S")
        split_timestamp = timestamp.split()
        x = [badge, split_timestamp[0], split_timestamp[1]]
        date_times.append(x)

# # Push all info in date_times list to database
def clock_times_collector():
    for dt in date_times:
        con = sqlite3.connect("test.db")
        c = con.cursor()
        c.execute("""CREATE TABLE IF NOT EXISTS clockTime (badge TEXT, date TEXT, time TEXT)""")

        query = """INSERT INTO clockTime (badge, date, time) VALUES (?, ?, ?)"""

        c.execute(query, (dt[0], dt[1], dt[2]))

        con.commit()

    con.close()

# clock_times_collector()

# ###############################################################################################
# COPY CLOCK TIMES TO WAGES TIME SHEET
# ###############################################################################################

# Connect to database
con = sqlite3.connect("test.db")
c = con.cursor()

# Load Wage Times workbook 
wb = load_workbook('Wage Times.xlsx')
ws = wb.active

# For each day copy actual clock in and out time
i = 0

for x in range(ws.max_row + 1):
    times_in_before = []
    times_in_late = []

    times_out_before = []
    times_out_late = []
    
    # Collect badge and date from excel sheet
    badge = ws.cell(row=2 + i, column=2).value
    date = ws.cell(row=2 + i, column=4).value

    # badge = '86'
    # date = '28/09/23'
    
    # loop through each employee and find actual clock in for day
    if badge != None and date != None:
        c.execute('SELECT time FROM clockTime WHERE badge = ? AND date = ?', (badge, date))
        clock_times = c.fetchall()
       
        # Convert tuple to list 
        clock_times_list = []
        for x in clock_times:
            clock_times_list.append(x[0])

        # Copy actual clock times to excel 
        for x in clock_times_list:
            ti = ws.cell(row=2 + i, column=5).value
            to = ws.cell(row=2 + i, column=6).value

            # Check clock in times vs roster
            if ti == 18:
                time_r = max(clock_times_list)
                t = time.fromisoformat(time_r).strftime('%H:%M')
                ws.cell(row=2+ i, column=7, value=t)
            elif ti == 0:
                ws.cell(row=2 + i, column=7, value='')
            else:
                time_r = min(clock_times_list)
                t = time.fromisoformat(time_r).strftime('%H:%M')
                ws.cell(row=2 + i, column=7, value=t)

            # Check clock out times vs roster
            if to == 6:
                time_r = min(clock_times_list)
                t = time.fromisoformat(time_r).strftime('%H:%M')
                ws.cell(row=2 + i, column=8, value=t)
            elif to == 0:
                ws.cell(row=2 + i, column=8, value='')
            else:
                time_r = max(clock_times_list)
                t = time.fromisoformat(time_r).strftime('%H:%M')
                ws.cell(row=2 + i, column=8, value=t)
    i += 1

wb.save('Wage Times.xlsx')
wb.close()