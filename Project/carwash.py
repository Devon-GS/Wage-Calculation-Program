from openpyxl import load_workbook
import sqlite3

# Load workbook carwash times
wb = load_workbook("Carwash Times/Carwash Times.xlsx", data_only=True)
ws = wb['Times']

rows = ws.iter_rows(min_row=3, max_row=9, min_col=12, max_col=16)

# Put all info in lis
carwash_times = []

for row in rows:
    if row[0].value != None:
        x = [str(row[0].value), str(row[1].value), str(row[2].value), str(row[3].value), '0']
        # x = [row[0].value, row[1].value, row[2].value, row[3].value, '0']
        carwash_times.append(x)


# Connect to database
con = sqlite3.connect("wageTimes.db")
c = con.cursor()
# Create table
c.execute(
    """CREATE TABLE IF NOT EXISTS carwashTotal (
        name TEXT,
        badge TEXT,
        normal TEXT,
        sunday TEXT,
        public TEXT
        )"""
)

# Add week one data to table
query = """INSERT INTO carwashTotal (
        name,
        badge,
        normal,
        sunday,
        public
        )
        VALUES (?, ?, ?, ?, ?)"""

# Add times to database
i = 0
for x in carwash_times:
    c.execute(query, (carwash_times[i][0], carwash_times[i][1], carwash_times[i][2], carwash_times[i][3], carwash_times[i][4]))
    i += 1

con.commit()
con.close()