import os
import xlwings as xw
from pathlib import Path
from openpyxl import Workbook 
from openpyxl.styles import Alignment, Font, Border, Side,  NamedStyle, PatternFill

# --- GET FILES WITH DYNAMIC NAMES ---
# Get payroll and carwash hours
def DYNAMIC_FILE_LOC(section):
	cwd = Path(__file__).parent 

	if section == 'Payroll':
		excel_file = cwd / "Payroll"
	else:
		excel_file = cwd / "Carwash Times" / "Carwash Hours"

	folder = Path(excel_file)
	
	# Search for Excel files and filter out hidden/temp files
	excel_files = folder.glob("*.xlsx*")
	valid_files =[file for file in excel_files if not file.name.startswith("~$")]
	
	# Check only one file in folder 
	if len(valid_files) == 1:
		return str(valid_files[0].resolve())	
	elif len(valid_files) == 0:
		return None
	else: 
		return None

# --- GET ALL STATIC DIRECTORIES --
DB_PATH = "wageTimes.db"
WAGE_TIMES_FILE = "Wage Times.xlsx"


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

BAKER_CASHIER_FILE = os.path.join(BASE_DIR, "Baker Cashier", "Baker Cashier Work.xlsx")
BADGE_NUMBER_FILE = os.path.join(BASE_DIR, "Badge Numbers", "Badges.xlsx")
PUBLIC_HOILIDAY_FILE = os.path.join(BASE_DIR, "Public Holidays", "Public Holidays.xlsx")
ROSTER_FOLDER = os.path.join(BASE_DIR, "Rosters")
UNICLOX_FOLDER = os.path.join(BASE_DIR, "Uniclox")
PUBLIC_HIOLIDAY_FILE = os.path.join(BASE_DIR, "Public Holidays", "Public Holidays.xlsx")
ATT_ROSTER_FILE = os.path.join(ROSTER_FOLDER, "Attendant_Carwash_Roster.xlsx")
CAS_ROSTER_FILE = os.path.join(ROSTER_FOLDER, "CASHIERS_ROSTER.xlsx")
CARWASH_FOLDER = os.path.join(BASE_DIR, "Carwash Times")
CARWASH_FILE = os.path.join(BASE_DIR, "Carwash Times", "Carwash Times.xlsx")
CARWASH_HOURS_FILE = os.path.join(BASE_DIR, "Carwash Times", "Carwash Hours", "Carwash Hours.xlsx")
TAX_RATES_FILE = os.path.join(BASE_DIR, "Tax", "Tax_rates", "PAYE_Fortnight.xlsx")
TAX_RESULTS = os.path.join(BASE_DIR, "Tax", "tax_results.xlsx")
PAYROLL_FOLDER = os.path.join(BASE_DIR, "Payroll")
PAYSLIP_TEMPLATE = os.path.join(BASE_DIR, "Templates", "Payslip_Template.xlsx")
PAYSLIP_FOLDER = os.path.join(BASE_DIR, "Payslips")
COPY_FOLDER = os.path.join(BASE_DIR, "Copy Folder")

# --- STYLES FOR EXCEL ---
THIN_SIDE = Side(style='thin', color="000000")
THICK_SIDE = Side(style='thick', color='000000')
DOUBLE_SIDE = Side(style='double', color="000000")

LEFT_ALIGN = Alignment(horizontal='left')
CENTER_ALIGN = Alignment(horizontal='center')
BOLD_FONT = Font(bold=True, underline='single')
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

TOTAL_STYLE = NamedStyle(
	name = "total_style",
	font = Font(bold=True),
	border = Border(top=THIN_SIDE, bottom=DOUBLE_SIDE)	
)

COLUMN_WIDTHS_ATT = {'A':15.00, 'B':13.57, 'C':10.71, 'D':10.0, 'E':6.86, 'F':8.43, 'G':12.00, 'H':13.71,
					'I':5.43, 'J':12.43, 'K':11.29, 'L':8.00, 'M':12.39, 'N':14.83, 'O':13.61}
COLUMN_WIDTHS_TOTALS = {'A':11.00, 'B':17.57, 'C':17.43, 'D':23.71, 'E':11.11, 'F':12.39, 'G':14.83, 'H':13.61}

COL_DIFF = 0.78

# --- INITILIZE EXCEL WAGE TIMES WORKBOOK ---
def CREATE_EXCEL():
	if not os.path.isfile(WAGE_TIMES_FILE):
		wb = Workbook()

		# Add named style
		wb.add_named_style(TOTAL_STYLE)

		# Create the bold font style
		header_font = Font(bold=True, underline='single')

		# Remove and add sheets
		wb.remove(wb['Sheet'])
		sheet_list = ['Att Week One', 'Att Week Two', 'Att Total', 
				'Cashier Week One', 'Cashier Week Two', 'Cashier Total']
		
		for sheet in sheet_list:
			wb.create_sheet(sheet)
		
		# Create heading
		for ws in wb.worksheets:
			if ws in [wb['Att Total'], wb['Cashier Total']]:
				if ws == wb['Att Total']:
					ws["A1"] = 'Name'
					ws["B1"] = 'Total Normal Hours'
					ws["C1"] = 'Total Sunday Hours'
					ws["D1"] = 'Total Public Holiday Hours'
					ws["E1"] = 'No Clock'
				else:
					ws["A1"] = 'Name'
					ws["B1"] = 'Total Normal Hours'
					ws["C1"] = 'Total Sunday Hours'
					ws["D1"] = 'Total Public Holiday Hours'
					ws["E1"] = 'No Clock'
					ws["F1"] = "Cashier Hours"
					ws["G1"] = "C. Sunday Hours"
					ws["H1"] = "C. Public Hours"
			else:
				ws["A1"] = "Name"
				ws["B1"] = "Badge Number"
				ws["C1"] = "Week Day"
				ws["D1"] = "Date"
				ws["E1"] = "Time In"
				ws["F1"] = "Time Out"
				ws["G1"] = "Clock Time In"
				ws["H1"] = "Clock Time Out"
				ws["I1"] = "Hours"
				ws["J1"] = "Sunday Hours"
				ws["K1"] = "Public Hours"
				ws["L1"] = "No Clock"
				if ws in [wb['Cashier Week One'], wb['Cashier Week Two']]:
					ws["M1"] = 'Cashier Hours'
					ws["N1"] = 'C. Sunday Hours'
					ws["O1"] = 'C. Public Hours'

			# Bold headings
			for cell in ws[1]:
				cell.font = header_font

		wb.save(WAGE_TIMES_FILE)
		wb.close()

# --- INITILIZE EXCEL CARWASH TIMES WORKBOOK ---
def CREATE_CARWASH_TIMES():
	# 1. Stop if the file already exists
	if os.path.isfile(CARWASH_FILE):
		return

	wb = Workbook()
	ws = wb.active
	ws.title = 'Times'

	# --- HELPER FUNCTION FOR BORDERS ---
	def apply_box_borders(cell_range):
		"""Applies a thick outside border and thin inside borders to any range."""
		rows = ws[cell_range]
		for i, row in enumerate(rows):
			for j, cell in enumerate(row):
				top = THICK_SIDE if i == 0 else THIN_SIDE
				bottom = THICK_SIDE if i == len(rows) - 1 else THIN_SIDE
				left = THICK_SIDE if j == 0 else THIN_SIDE
				right = THICK_SIDE if j == len(row) - 1 else THIN_SIDE
				
				cell.border = Border(top=top, bottom=bottom, left=left, right=right)

	# 2. INSERT HEADERS
	# Week 1
	ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'] = 'Name', 'Badge', 'Day', 'Date', 'Hour'
	
	# Week 2
	ws['G1'], ws['H1'], ws['I1'], ws['J1'], ws['K1'] = 'Name', 'Badge', 'Day', 'Date', 'Hour'
	
	# Total Hours
	ws['M1'], ws['N1'], ws['O1'], ws['P1'] = 'Name', 'Badge Number', 'Total Normal', 'Total Sunday'
	
	# Extra Times
	ws['M11'] = 'EXTRA'
	ws['M12'], ws['N12'], ws['O12'], ws['P12'] = 'Name', 'Badge Number', 'Early Times', 'Amount'

	# 3. APPLY HEADER STYLES (Row 1)
	for col_num in range(1, 17):
		cell = ws.cell(row=1, column=col_num)
		cell.font = BOLD_FONT
		cell.alignment = CENTER_ALIGN

	# 4. SET COLUMN WIDTHS
	column_widths = {
		'A': 9.72, 'B': 8.83, 'C': 10.42, 'D': 10.52, 'E': 8.83,
		'G': 9.72, 'H': 8.83, 'I': 10.42, 'J': 10.52, 'K': 8.83,
		'M': 9.72, 'N': 12.90, 'O': 11.83, 'P': 11.94
	}
	for col_letter, width in column_widths.items():
		ws.column_dimensions[col_letter].width = width

	# Format Badge columns (B and H) to be left-aligned 
	# (Pre-formatting up to row 20)
	for row in range(2, 65): 
		ws[f'B{row}'].alignment = LEFT_ALIGN
		ws[f'H{row}'].alignment = LEFT_ALIGN

	# 5. MERGED CELLS & ALIGNMENT FOR "EXTRA TIMES"
	ws.merge_cells('M11:P11')
	
	# Apply center alignment (and bold) to rows 11 and 12
	for row in ws['M11:P12']:
		for cell in row:
			cell.alignment = CENTER_ALIGN
			cell.font = BOLD_FONT  

	# 6. APPLY YELLOW HIGHLIGHTS
	# Top grid highlights
	for row in ws['M2:P9']:
		for cell in row:
			cell.fill = YELLOW_FILL

			# Don't align coloumn M			
			if cell.column_letter != 'M':
				cell.alignment = CENTER_ALIGN

	# Highlight columns N and P for rows 13 through 20
	for row in ws['M13:P20']:
		for cell in row:
			# Don't align coloumn M	
			if cell.column_letter != 'M':
				cell.alignment = CENTER_ALIGN
			# Yellow fill
			if cell.column_letter in ('N', 'P'):
				cell.fill = YELLOW_FILL

	# 7. APPLY BORDERS
	apply_box_borders('M2:P9')
	apply_box_borders('M11:P11')
	apply_box_borders('M12:P20')

	# 8. ADD FORMULAS FOR EXTRA TIME
	# Loop through rows 13 to 20
	for row_num in range(13, 21):
		ws[f'P{row_num}'].value = f"=O{row_num}*50"

	# SAVE WORKBOOK
	wb.save(CARWASH_FILE)

# --- HELPER FUNCTIONS FOR EXCEL ---
def RECALCULATE_EXCEL_FORMULAS(filepath):
	"""Opens and saves an Excel file in the background to force formula recalculation."""
	with xw.App(visible=False) as app:
		wb = app.books.open(filepath)
		wb.save()
		wb.close()