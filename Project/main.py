from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import sqlite3

# ==============================================================================
# FUNCTIONS
# ==============================================================================

# Get Time in / Time Out
def first(weekday):
    if weekday == 'AF' or weekday == ' ' or weekday == '0':
        return 0.0
    else:
        first = float(re.findall('[0-9]+(?=.*\-)', weekday)[0])
        return first

def second(weekday):
    if weekday == 'AF' or weekday == ' ' or weekday == '' or weekday == '0':
        return 0.0
    else:
        second = float(re.findall('\-(.*)', weekday)[0])
        return second

# ==============================================================================
# IMPORT ROSTER TIMES AND DATES AND BADGE (ATTENDENTS)
# ==============================================================================
file = '../Attendant_Carwash_Roster.xls'

# Get Columns
columns = ['idx','ATTENDANTS', 'THURS', 'FRI', 'SAT', 'SUN', 'MON', 'TUE', 'WED']

# Get Times 
data = pd.read_excel(file, index_col=0, header=1, usecols=columns)
data = data.fillna(0)

# Get Dates
data_date = pd.read_excel(file, header=None, nrows=2, usecols='C:I')
data_date_ex = data_date.loc[0]

weekone_dates = {}
for x in data_date_ex:
	weekone_dates[x.strftime("%A")] = x.date().strftime("%d/%m/%y")

# Get week one from excel sheet
week_one_data = data.loc[0:14]
week_one = []
for x in week_one_data.to_numpy().tolist():
    if str(x[0]) != 'nan':
        if x[0] != 0:
            week_one.append(x)

# Get badges for badges.xlsx and convert to dict
badge_file = 'Badge Numbers/badges.xlsx'
bf = pd.read_excel(badge_file, header=None)
badge = bf.values.tolist()

badges = {}
for x, y in badge:
    badges[x] = y

# ==============================================================================
# CREATE DATABASE SQLITE
# ==============================================================================

def db_init():
    # Connect to database
    con = sqlite3.connect("weekOne.db")
    c = con.cursor()
    # Create table
    c.execute("""CREATE TABLE IF NOT EXISTS roster (
            name TEXT,
            badge TEXT,
            thur TEXT,
            fri TEXT,
            sat TEXT,
            sun TEXT,
            mon TEXT,
            tue TEXT,
            wed TEXT
            )""")

    # Add week one data to table
    week1 = """INSERT INTO roster (
            name,
            badge,
            thur,
            fri,
            sat,
            sun,
            mon,
            tue,
            wed
            )
            VALUES (?, ?, ?, ?, ? ,? ,?, ?, ?)"""

    c.execute(week1, ('Date', '999', 0, 0, 0, 0, 0, 0, 0))

    for week in week_one:
        x = (week[0], 0, week[1],week[2],week[3],week[4],week[5],week[6],week[7])
        c.execute(week1, (x))

        con.commit()
    con.close()

def db_update_dates():
    # Update table with roster dates
    con = sqlite3.connect("weekOne.db")
    c = con.cursor()

    query = ("""UPDATE roster SET
            thur = ?,
            fri = ?,
            sat = ?,
            sun = ?,
            mon = ?,
            tue = ?,
            wed = ?
            WHERE badge = ?
            """)
    thursday = weekone_dates['Thursday']
    friday = weekone_dates['Friday']
    saturday = weekone_dates['Saturday']
    sunday = weekone_dates['Sunday']
    monday = weekone_dates['Monday']
    tuesday = weekone_dates['Tuesday']
    wednesday = weekone_dates['Wednesday']

    c.execute(query, (thursday, friday, saturday, sunday, monday, tuesday, wednesday, 999))
    con.commit()
    con.close()

def db_update_badges():
    # Update table with badge numbers
    con = sqlite3.connect("weekOne.db")
    c = con.cursor()

    query = ("""UPDATE roster SET badge = ? WHERE name = ?""")

    for x in badges:
        c.execute(query, (badges[x], x))
        con.commit()

    con.close()


# db_init()
# db_update_dates()
# db_update_badges()



# Grab data from table for excel workbook
con = sqlite3.connect("weekOne.db")
c = con.cursor()

c.execute('SELECT name FROM roster')
name_records = c.fetchall()

week_one_data = []

for record in name_records:
    # Grab data from database using name of person
    c.execute('SELECT * FROM roster where name=?', (record[0],))
    r = c.fetchall()
    week_one_data.append(r)

con.close()


# ==============================================================================
# TEST CODE START
# ==============================================================================



# ==============================================================================
# TEST CODE END
# ==============================================================================

# ==============================================================================
# CREATE TIME EXCEL FILE
# ==============================================================================

wb = Workbook()
ws = wb.active
ws.title = 'Wages Calculator'

# Create Column Headings
ws['A1'] = 'Name'
ws['B1'] = 'Badge Number'
ws['C1'] = 'Date'
ws['D1'] = 'Week Day'
ws['E1'] = 'Time In'
ws['F1'] = 'Time Out'
ws['G1'] = 'Hours'

# Get Time in / Time Out
# Start row and col
t_row = 2
t_col = 1
# Date and day start row and col
d_row = 2
d_col = 3
# Badge number start row nnd col
b_row = 2
b_col = 2
# Date and day continue
i_row_d = 0
# Name contine row
i_row_n = 0
# badge row contine
i_row_b = 0
# Row continue
i_row = 0 

# Loop through database and save to excel sheet
for r in week_one_data[1:]:
    name = r[0][0]
    badge = r[0][1]
    thur = r[0][2]
    fri = r[0][3]
    sat = r[0][4]
    sun = r[0][5]
    mon = r[0][6]
    tue = r[0][7]
    wed = r[0][8]

    # Get day and date from dict
    thursday = weekone_dates['Thursday']
    friday = weekone_dates['Friday']
    saturday = weekone_dates['Saturday']
    sunday = weekone_dates['Sunday']
    monday = weekone_dates['Monday']
    tuesday = weekone_dates['Tuesday']
    wednesday = weekone_dates['Wednesday']
    
    # Thursday
    if thur == 'AF':
        thur = 0
        # Add name to column A
        ws.cell(t_row + i_row_n, t_col, value=name)
        # Add roster time to column E and F
        ws.cell(t_row + i_row, t_col + 4, value=thur)
        ws.cell(t_row + i_row, t_col + 5, value=thur)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Thursday')
        ws.cell(d_row + i_row_d, d_col + 1, value=thursday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
    
    elif first(thur) == 18:
        thur_s = first(thur)
        thur_e = second(thur)
        # Add name to column A and start time to column E
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=thur_s)
        # Add name to column A and shift row down by one and
        # Add end time to column F
        ws.cell(t_row + i_row_n + 1, t_col, value=name)
        ws.cell(t_row + i_row + 1, t_col + 5, value=thur_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Thursday')
        ws.cell(d_row + i_row_d, d_col + 1, value=thursday)
        # Add dates to columns C and D and shift row down by 1
        ws.cell(d_row + i_row_d + 1, d_col, value='Thursday')
        ws.cell(d_row + i_row_d + 1, d_col + 1, value=thursday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
        ws.cell(b_row + i_row_b + 1, b_col, value=badge)
        # Add one to shift rows down for next days info
        i_row_n += 1
        i_row += 1
        i_row_d += 1
        i_row_b += 1
    else:
        thur_s = first(thur)
        thur_e = second(thur)
        # Add name to column A and roster times to column E and F
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=thur_s)
        ws.cell(t_row + i_row, t_col + 5, value=thur_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Thursday')
        ws.cell(d_row + i_row_d, d_col + 1, value=thursday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)

    # Add one to shift rows down for next day info
    i_row_n += 1
    i_row += 1
    i_row_d += 1 
    i_row_b += 1
    
    # Friday
    if fri == 'AF':
        fri = 0
        # Add name to column A
        ws.cell(t_row + i_row_n, t_col, value=name)
        # Add roster time to column E and F
        ws.cell(t_row + i_row, t_col + 4, value=fri)
        ws.cell(t_row + i_row, t_col + 5, value=fri)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Friday')
        ws.cell(d_row + i_row_d, d_col + 1, value=friday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
    
    elif first(fri) == 18:
        fri_s = first(fri)
        fri_e = second(fri)
        # Add name to column A and start time to column E
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=fri_s)
        # Add name to column A and shift row down by one and
        # Add end time to column F
        ws.cell(t_row + i_row_n + 1, t_col, value=name)
        ws.cell(t_row + i_row + 1, t_col + 5, value=fri_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Friday')
        ws.cell(d_row + i_row_d, d_col + 1, value=friday)
        # Add dates to columns C and D and shift row down by 1
        ws.cell(d_row + i_row_d + 1, d_col, value='Friday')
        ws.cell(d_row + i_row_d + 1, d_col + 1, value=friday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
        ws.cell(b_row + i_row_b + 1, b_col, value=badge)
        # Add one to shift rows down for next days info
        i_row_n += 1
        i_row += 1
        i_row_d += 1
        i_row_b += 1
    else:
        fri_s = first(fri)
        fri_e = second(fri)
        # Add name to column A and roster times to column E and F
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=fri_s)
        ws.cell(t_row + i_row, t_col + 5, value=fri_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Friday')
        ws.cell(d_row + i_row_d, d_col + 1, value=friday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)

    # Add one to shift rows down for next dayd info
    i_row_n += 1
    i_row += 1
    i_row_d += 1 
    i_row_b += 1

    # Saturday
    if sat == 'AF':
        sat = 0
        # Add name to column A
        ws.cell(t_row + i_row_n, t_col, value=name)
        # Add roster time to column E and F
        ws.cell(t_row + i_row, t_col + 4, value=sat)
        ws.cell(t_row + i_row, t_col + 5, value=sat)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Saturday')
        ws.cell(d_row + i_row_d, d_col + 1, value=saturday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
    
    elif first(sat) == 18:
        sat_s = first(sat)
        sat_e = second(sat)
        # Add name to column A and start time to column E
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=sat_s)
        # Add name to column A and shift row down by one and
        # Add end time to column F
        ws.cell(t_row + i_row_n + 1, t_col, value=name)
        ws.cell(t_row + i_row + 1, t_col + 5, value=sat_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Saturday')
        ws.cell(d_row + i_row_d, d_col + 1, value=saturday)
        # Add dates to columns C and D and shift row down by 1
        ws.cell(d_row + i_row_d + 1, d_col, value='Saturday')
        ws.cell(d_row + i_row_d + 1, d_col + 1, value=saturday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
        ws.cell(b_row + i_row_b + 1, b_col, value=badge)
        # Add one to shift rows down for next days info
        i_row_n += 1
        i_row += 1
        i_row_d += 1
        i_row_b += 1
    else:
        sat_s = first(sat)
        sat_e = second(sat)
        # Add name to column A and roster times to column E and F
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=sat_s)
        ws.cell(t_row + i_row, t_col + 5, value=sat_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Saturday')
        ws.cell(d_row + i_row_d, d_col + 1, value=saturday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
    
    # Add one to shift rows down for next dayd info
    i_row_n += 1
    i_row += 1
    i_row_d += 1 
    i_row_b += 1

    # Sunday
    if sun == 'AF':
        sun = 0
        # Add name to column A
        ws.cell(t_row + i_row_n, t_col, value=name)
        # Add roster time to column E and F
        ws.cell(t_row + i_row, t_col + 4, value=sun)
        ws.cell(t_row + i_row, t_col + 5, value=sun)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Sunday')
        ws.cell(d_row + i_row_d, d_col + 1, value=sunday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
    
    elif first(sun) == 18:
        sun_s = first(sun)
        sun_e = second(sun)
        # Add name to column A and start time to column E
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=sun_s)
        # Add name to column A and shift row down by one and
        # Add end time to column F
        ws.cell(t_row + i_row_n + 1, t_col, value=name)
        ws.cell(t_row + i_row + 1, t_col + 5, value=sun_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Sunday')
        ws.cell(d_row + i_row_d, d_col + 1, value=sunday)
        # Add dates to columns C and D and shift row down by 1
        ws.cell(d_row + i_row_d + 1, d_col, value='Sunday')
        ws.cell(d_row + i_row_d + 1, d_col + 1, value=sunday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
        ws.cell(b_row + i_row_b + 1, b_col, value=badge)
        # Add one to shift rows down for next days info
        i_row_n += 1
        i_row += 1
        i_row_d += 1
        i_row_b += 1
    else:
        sun_s = first(sun)
        sun_e = second(sun)
        # Add name to column A and roster times to column E and F
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=sun_s)
        ws.cell(t_row + i_row, t_col + 5, value=sun_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Sunday')
        ws.cell(d_row + i_row_d, d_col + 1, value=sunday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
    
    # Add one to shift rows down for next dayd info
    i_row_n += 1
    i_row += 1
    i_row_d += 1
    i_row_b += 1

    # Monday
    if mon == 'AF':
        mon = 0
        # Add name to column A
        ws.cell(t_row + i_row_n, t_col, value=name)
        # Add roster time to column E and F
        ws.cell(t_row + i_row, t_col + 4, value=mon)
        ws.cell(t_row + i_row, t_col + 5, value=mon)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Monday')
        ws.cell(d_row + i_row_d, d_col + 1, value=monday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
    
    elif first(mon) == 18:
        mon_s = first(mon)
        mon_e = second(mon)
        # Add name to column A and start time to column E
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=mon_s)
        # Add name to column A and shift row down by one and
        # Add end time to column F
        ws.cell(t_row + i_row_n + 1, t_col, value=name)
        ws.cell(t_row + i_row + 1, t_col + 5, value=mon_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Monday')
        ws.cell(d_row + i_row_d, d_col + 1, value=monday)
        # Add dates to columns C and D and shift row down by 1
        ws.cell(d_row + i_row_d + 1, d_col, value='Monday')
        ws.cell(d_row + i_row_d + 1, d_col + 1, value=monday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
        ws.cell(b_row + i_row_b + 1, b_col, value=badge)
        # Add one to shift rows down for next days info
        i_row_n += 1
        i_row += 1
        i_row_d += 1
        i_row_b += 1
    else:
        mon_s = first(mon)
        mon_e = second(mon)
        # Add name to column A and roster times to column E and F
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=mon_s)
        ws.cell(t_row + i_row, t_col + 5, value=mon_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Monday')
        ws.cell(d_row + i_row_d, d_col + 1, value=monday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
    
    # Add one to shift rows down for next dayd info
    i_row_n += 1
    i_row += 1
    i_row_d += 1
    i_row_b += 1

    # Tueday
    if tue == 'AF':
        tue = 0
        # Add name to column A
        ws.cell(t_row + i_row_n, t_col, value=name)
        # Add roster time to column E and F
        ws.cell(t_row + i_row, t_col + 4, value=tue)
        ws.cell(t_row + i_row, t_col + 5, value=tue)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Tuesday')
        ws.cell(d_row + i_row_d, d_col + 1, value=tuesday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
    
    elif first(tue) == 18:
        tue_s = first(tue)
        tue_e = second(tue)
        # Add name to column A and start time to column E
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=tue_s)
        # Add name to column A and shift row down by one and
        # Add end time to column F
        ws.cell(t_row + i_row_n + 1, t_col, value=name)
        ws.cell(t_row + i_row + 1, t_col + 5, value=tue_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Tuesday')
        ws.cell(d_row + i_row_d, d_col + 1, value=tuesday)
        # Add dates to columns C and D and shift row down by 1
        ws.cell(d_row + i_row_d + 1, d_col, value='Tuesday')
        ws.cell(d_row + i_row_d + 1, d_col + 1, value=tuesday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
        ws.cell(b_row + i_row_b + 1, b_col, value=badge)
        # Add one to shift rows down for next days info
        i_row_n += 1
        i_row += 1
        i_row_d += 1
        i_row_b += 1
    else:
        tue_s = first(tue)
        tue_e = second(tue)
        # Add name to column A and roster times to column E and F
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=tue_s)
        ws.cell(t_row + i_row, t_col + 5, value=tue_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Tuesday')
        ws.cell(d_row + i_row_d, d_col + 1, value=tuesday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
    
    # Add one to shift rows down for next dayd info
    i_row_n += 1
    i_row += 1
    i_row_d += 1
    i_row_b += 1

    # Wednesday
    if wed == 'AF':
        wed = 0
        # Add name to column A
        ws.cell(t_row + i_row_n, t_col, value=name)
        # Add roster time to column E and F
        ws.cell(t_row + i_row, t_col + 4, value=wed)
        ws.cell(t_row + i_row, t_col + 5, value=wed)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Wednesday')
        ws.cell(d_row + i_row_d, d_col + 1, value=wednesday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
    
    elif first(wed) == 18:
        wed_s = first(wed)
        wed_e = second(wed)
        # Add name to column A and start time to column E
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=wed_s)
        # Add name to column A and shift row down by one and
        # Add end time to column F
        ws.cell(t_row + i_row_n + 1, t_col, value=name)
        ws.cell(t_row + i_row + 1, t_col + 5, value=wed_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Wednesday')
        ws.cell(d_row + i_row_d, d_col + 1, value=wednesday)
        # Add dates to columns C and D and shift row down by 1
        ws.cell(d_row + i_row_d + 1, d_col, value='Wednesday')
        ws.cell(d_row + i_row_d + 1, d_col + 1, value=wednesday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
        ws.cell(b_row + i_row_b + 1, b_col, value=badge)
        # Add one to shift rows down for next days info
        i_row_n += 1
        i_row += 1
        i_row_d += 1
        i_row_b += 1
    else:
        wed_s = first(wed)
        wed_e = second(wed)
        # Add name to column A and roster times to column E and F
        ws.cell(t_row + i_row_n, t_col, value=name)
        ws.cell(t_row + i_row, t_col + 4, value=wed_s)
        ws.cell(t_row + i_row, t_col + 5, value=wed_e)
        # Add dates to columns C and D
        ws.cell(d_row + i_row_d, d_col, value='Wednesday')
        ws.cell(d_row + i_row_d, d_col + 1, value=wednesday)
        # Add badge number to column B
        ws.cell(b_row + i_row_b, b_col, value=badge)
    
    # Add one to shift rows down for next dayd info
    i_row_n += 1
    i_row += 1
    i_row_d += 1
    i_row_b += 1 
    
    
    # Add one to shift rows down for next dayd info
    i_row_n += 2
    i_row += 2
    i_row_d += 2
    i_row_b += 2
    con.close()

# Close workbook
wb.save('Wage Times.xlsx')
wb.close()

# wb = load_workbook('Wage Times.xlsx')
# ws = wb.active