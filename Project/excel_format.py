from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side

# Column letters and widths
columns_att = {'A':15.00, 'B':12.33, 'C':9.7, 'D':7.8, 'E':6.3, 'F':7.7, 'G':11.11, 'H':12.70, 'I':5.30, 'J':11.30, 'K':10.30, 'L':7.60, 'M':11.70}
columns_total = {'A':11.00, 'B':16.44, 'C':16.00, 'D':21.78, 'E':11.11, 'F':17.11}

col_diff = 0.78

def cell_center(sheet):
    i = 0
    for x in range(sheet.max_row):
        sheet.cell(row=2 + i, column=2).alignment = Alignment(horizontal='center')
        sheet.cell(row=2 + i, column=3).alignment = Alignment(horizontal='center')
        sheet.cell(row=2 + i, column=4).alignment = Alignment(horizontal='center')
        sheet.cell(row=2 + i, column=5).alignment = Alignment(horizontal='center')
        sheet.cell(row=2 + i, column=6).alignment = Alignment(horizontal='center')
        i += 1

def excel_format():
    wb = load_workbook("Wage Times.xlsx")

    ws = wb['Att Week One']
    wso = wb['Att Week Two']
    wst = wb['Att Total']

    wsc = wb['Cashier Week One']
    wsco = wb['Cashier Week Two']
    wsct = wb['Cashier Total']
    
    # Format Attendents
    for col, size in columns_att.items():
        ws.column_dimensions[col].width = size + col_diff
        wso.column_dimensions[col].width = size + col_diff    

    for col, size in columns_total.items():
        wst.column_dimensions[col].width = size + col_diff

    cell_center(wst)

    # Format Cashiers
    for col, size in columns_att.items():
        wsc.column_dimensions[col].width = size + col_diff
        wsco.column_dimensions[col].width = size + col_diff

    for col, size in columns_total.items():
        wsct.column_dimensions[col].width = size + col_diff

    cell_center(wsct)

    # Format total rows to bold and border
    total_format = NamedStyle(name="total_format")
    total_format.font = Font(bold=True)
    bt = Side(style='thin', color="000000") 
    bb = Side(style='double', color="000000") 
    total_format.border = Border(top=bt, bottom=bb) 

    # Apply styles att weekone
    i = 0   
    for x in range(ws.max_row + 1):
        name = ws.cell(row=2 + i, column=1).value 
        if name != None:
            if 'Total' in name:
                ws.cell(row=2 + i, column=1).style = total_format
                ws.cell(row=2 + i, column=2).style = total_format
                ws.cell(row=2 + i, column=9).style = total_format
                ws.cell(row=2 + i, column=10).style = total_format
                ws.cell(row=2 + i, column=11).style = total_format
                ws.cell(row=2 + i, column=12).style = total_format
        i += 1

    # Apply styles att weektwo
    i = 0   
    for x in range(wso.max_row + 1):
        name = wso.cell(row=2 + i, column=1).value 
        if name != None:
            if 'Total' in name:
                wso.cell(row=2 + i, column=1).style = total_format
                wso.cell(row=2 + i, column=2).style = total_format
                wso.cell(row=2 + i, column=9).style = total_format
                wso.cell(row=2 + i, column=10).style = total_format
                wso.cell(row=2 + i, column=11).style = total_format
                wso.cell(row=2 + i, column=12).style = total_format
        i += 1
    
    # Apply styles cashier weekone
    i = 0   
    for x in range(wsc.max_row + 1):
        name = wsc.cell(row=2 + i, column=1).value 
        if name != None:
            if 'Total' in name:
                wsc.cell(row=2 + i, column=1).style = total_format
                wsc.cell(row=2 + i, column=2).style = total_format
                wsc.cell(row=2 + i, column=9).style = total_format
                wsc.cell(row=2 + i, column=10).style = total_format
                wsc.cell(row=2 + i, column=11).style = total_format
                wsc.cell(row=2 + i, column=12).style = total_format
        i += 1
    
    # Apply styles cashier weektwo
    i = 0   
    for x in range(wsco.max_row + 1):
        name = wsco.cell(row=2 + i, column=1).value 
        if name != None:
            if 'Total' in name:
                wsco.cell(row=2 + i, column=1).style = total_format
                wsco.cell(row=2 + i, column=2).style = total_format
                wsco.cell(row=2 + i, column=9).style = total_format
                wsco.cell(row=2 + i, column=10).style = total_format
                wsco.cell(row=2 + i, column=11).style = total_format
                wsco.cell(row=2 + i, column=12).style = total_format
        i += 1

    wb.save("Wage Times.xlsx")
    wb.close()