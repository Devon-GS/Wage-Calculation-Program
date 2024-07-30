import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter import *
from tkinter.filedialog import askopenfilename
import pandas as pd
import os

# FUNCTIONS
# Function to clean tax anount
def clean_tax(x):
    x = x.replace(',', '')
    x = x.replace('R', '')
    x = x.replace(' ', '')
    if type(x) == str:
        x = int(x)
    return x

# Function to clean commas and currency sign from columns
def clean_column(x):
    x = x.replace(',', '')
    x = x.replace('R', '')
    return x

# MAIN PROGRAM CALCULATE TAX
def tax():
    # Run and save excel in background to reevaluate formula
    app = xw.App(visible=False)
    book = app.books.open("Payroll/Payroll.xlsx")
    # sheet = book.sheets[worksheet_name]
    # print(sheet['C5'].value)
    book.save() 
    book.close()
    app.quit()
    
    #  Read in tax deduction and weekly wage excel file
    df = pd.read_excel('Tax/Tax_rates/PAYE_Fortnight.xlsx')
    # file = askopenfilename(title='Select Wage File',initialdir=os.getcwd(), filetypes =[('xls', 'xlsx')])
    file = 'Payroll/Payroll.xlsx'
    df2 = pd.read_excel(file)

    # Clean columns Remuneration 1 and Remuneration 2 and change data type to int
    df['Remuneration 1'] = df['Remuneration 1'].apply(clean_column).astype(int)
    df['Remuneration 2'] = df['Remuneration 2'].apply(clean_column).astype(int)

    # Collect employee name and gross wage and place in dict
    employee_unfilted = []
    for x in df2.columns:
        employee_unfilted.append(x) 
        
    employee = employee_unfilted[2:-1]

    gross_wages_unfilted = []
    for x in df2.loc[19]:
        gross_wages_unfilted.append(x)

    gross_wage = gross_wages_unfilted[2:-1]

    d = dict(zip(employee, gross_wage))

    # Round values to 2 decimal places
    for key, value in d.items():
        d[key] = round(value, 2)

    # Merge any dublicate and drop dublicates
    for x in d.copy():
        if x[-1] == '1':
            dub = x[:-2]
            amt = d[dub] + d[x]
            d[dub] = amt
            del d[x]

    # Function to calculate tax payable
    def calculate_tax(gross_wage, tax_brackets):
        tax_payable = 0
        
        for i in range(len(tax_brackets)):
            min_income = tax_brackets['Remuneration 1'][i]
            max_income = tax_brackets['Remuneration 2'][i]
            tax_amount = tax_brackets['Under 65'][i]

            if min_income <= gross_wage <= max_income:
                tax = clean_tax(tax_amount)
                tax_payable = tax
                break

        return tax_payable

    # Create a new dictionary with name, gross wage, and tax payable for each person
    results = {}

    for name, gross_wage in d.items():
        tax = calculate_tax(gross_wage, df)
        results[name] = {'Gross Wage': gross_wage, 'Tax Payable': tax}

    # CREATE EXCEL SHEET WITH INFO FROM DICT

    # Create a DataFrame from the results dictionary
    results_df = pd.DataFrame.from_dict(results, orient='index')

    # Transpose the DataFrame
    results_df = results_df.T

    # Specify the path and filename for the Excel file
    output_file = 'Tax/tax_results.xlsx'

    # Check if the output file already exists, and remove it if it does
    if os.path.isfile(output_file):
        os.remove(output_file)

    # Write the DataFrame to Excel
    results_df.to_excel(output_file, sheet_name='Results', index=False)

    # Open the workbook
    wb = load_workbook(output_file)

    # Get the worksheet
    worksheet = wb['Results']

    # Set the labels in the first column
    labels = ['Employee Name', 'Gross Wage', 'Tax Payable']

    for i, label in enumerate(labels):
        worksheet.cell(row=i+1, column=1, value=label)

    # Set the column headers
    column_headers = results_df.columns
    for i, header in enumerate(column_headers):
        worksheet.cell(row=1, column=i+2, value=header)

    # Write the gross wages and tax payable rows
    gross_wages = results_df.loc['Gross Wage']
    tax_payable = results_df.loc['Tax Payable']

    for i, value in enumerate(gross_wages):
        worksheet.cell(row=2, column=i+2, value=value)

    for i, value in enumerate(tax_payable):
        worksheet.cell(row=3, column=i+2, value=value)

    # Save the Excel file
    wb.save(output_file)

    # ######################################################
    # Update Payroll File
    # ######################################################

    wb = load_workbook("Payroll/Payroll.xlsx")
    ws = wb.active

    # (dot = data_only True). To get uif amounts with out formula
    wb_dot = load_workbook("Payroll/Payroll.xlsx", data_only=True)
    ws_dot = wb_dot.active

    for col in range(3,ws_dot.max_column):
        col_letter = get_column_letter(col)
        name = ws_dot[f'{col_letter}1'].value
        uif = ws_dot[f'{col_letter}22'].value

        if uif != None:
            if uif > 0:
                tax_amt = results[name]['Tax Payable']
                ws[f'{col_letter}29'] = tax_amt

    wb.save("Payroll/Payroll.xlsx")
    wb.close()

    # Run and save excel in background to reevaluate formula
    app = xw.App(visible=False)
    book = app.books.open("Payroll/Payroll.xlsx")
    # sheet = book.sheets[worksheet_name]
    # print(sheet['C5'].value)
    book.save() 
    book.close()
    app.quit()
