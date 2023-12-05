# ###############################################################################################
# CASHIERS WEEK 1 - CALCULATE ROSTER VS CLOCK TIME IN EXCEL
# ###############################################################################################
import os
from datetime import datetime, time
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import sqlite3

# Load public holidays date
wb = load_workbook("Public Holidays/Public Holidays.xlsx", data_only=True)
ws = wb.active

public_holidays = []
for row in ws.iter_rows(min_row=2, max_col=1, max_row=20, values_only=True):
    x = row[0]
    if x != None:
        holiday_date = f'{x.day}/{x.month}/{str(x.year)[-2:]}'
        dateobj = datetime.strptime(holiday_date, '%d/%m/%y').date().strftime('%d/%m/%y')
        public_holidays.append(dateobj)

wb.close()

# Load baker cashier hours
wb = load_workbook("Baker Cashier/Baker Cashier Work.xlsx", data_only=True)
ws = wb.active

bc_working = []
for row in ws.iter_rows(min_row=2, max_col=2, max_row=20, values_only=True):
    x = row
    if x[0] != None:
        name = x[0]
        cashier_date = f'{x[1].day}/{x[1].month}/{str(x[1].year)[-2:]}'
        dateobj = datetime.strptime(cashier_date, '%d/%m/%y').date().strftime('%d/%m/%y')
        bc = [name, dateobj]
        bc_working.append(bc)

wb.close()

# ==============================================================================
# CALCULATE CLOCK IN AND CLOCK OUT TIMES WEEK ONE
# ==============================================================================

def cas_times_weekone():
    wb = load_workbook("Wage Times.xlsx")
    ws = wb['Cashier Week One']

    i = 0
    for x in range(ws.max_row + 1):
        tti = 0
        tto = 0

        name = ws.cell(row=2 + i, column=1).value

        day = ws.cell(row=2 + i, column=3).value

        time_in = ws.cell(row=2 + i, column=5).value
        clock_in = ws.cell(row=2 + i, column=7).value

        time_out = ws.cell(row=2 + i, column=6).value
        clock_out = ws.cell(row=2 + i, column=8).value

        if name == None:
            i += 1

        elif time_in > 0 and clock_in == None or time_out > 0 and clock_out == None:
            hours = 'No Clock'
            ws.cell(row=2 + i, column=12, value=hours)

            i += 1

        elif day == "Sunday":
            if time_in == 18:
                ti = time(time_in).strftime("%H:%M")
                ci = clock_in

                if ci > ti:
                    if ci[-2:] <= "05":
                        tti = 24.0 - float(ci[0:2])
                    elif ci[-2:] <= "15":
                        tti = 24.0 - (float(ci[0:2]) + 0.25)
                    elif ci[-2:] <= "30":
                        tti = 24.0 - (float(ci[0:2]) + 0.50)
                    elif ci[-2:] <= "45":
                        tti = 24.0 - (float(ci[0:2]) + 0.75)
                    elif ci[-2:] >= "45":
                        tti = 24.0 - (float(ci[0:2]) + 1.0)
                else:
                    tti = 24.0 - float(time_in)

                hours = tti
                ws.cell(row=2 + i, column=10, value=hours)

            elif clock_in == None and clock_out == None:
                hours = 0
                ws.cell(row=2 + i, column=10, value=hours)

            elif clock_in == None:
                # Calculate time out
                to = time(time_out).strftime("%H:%M")
                co = clock_out

                if co < to:
                    if co[-2:] <= "05":
                        tto = float(co[0:2]) + 0.00
                    elif co[-2:] <= "15":
                        tto = float(co[0:2]) + 0.25
                    elif co[-2:] <= "30":
                        tto = float(co[0:2]) + 0.50
                    elif co[-2:] <= "45":
                        tto = float(co[0:2]) + 0.75
                    elif co[-2:] >= "45":
                        tto = (float(co[0:2]) + 1) + 0.00
                else:
                    tto = int(time_out)

                hours = float(tto)
                ws.cell(row=2 + i, column=10, value=hours)

            else:
                ti = time(time_in).strftime("%H:%M")
                ci = clock_in

                to = time(time_out).strftime("%H:%M")
                co = clock_out

                # Calculate time in
                if ci > ti:
                    if ci[-2:] <= "05":
                        tti = float(ci[0:2])
                    elif ci[-2:] <= "15":
                        tti = float(ci[0:2]) + 0.25
                    elif ci[-2:] <= "30":
                        tti = float(ci[0:2]) + 0.50
                    elif ci[-2:] <= "45":
                        tti = float(ci[0:2]) + 0.75
                    elif ci[-2:] >= "45":
                        tti = (float(ci[0:2]) + 1) + 0.00
                else:
                    tti = float(time_in)

                # Calculate time out
                if co < to:
                    if co[-2:] <= "05":
                        tto = float(co[0:2]) + 0.00
                    elif co[-2:] <= "15":
                        tto = float(co[0:2]) + 0.25
                    elif co[-2:] <= "30":
                        tto = float(co[0:2]) + 0.50
                    elif co[-2:] <= "45":
                        tto = float(co[0:2]) + 0.75
                    elif co[-2:] >= "45":
                        tto = (float(co[0:2]) + 1) + 0.00
                else:
                    tto = int(time_out)

                # calculate hours worked
                hours = float(tto) - float(tti)
                ws.cell(row=2 + i, column=10, value=hours)

            i += 1

        elif clock_in == None and clock_out == None:
            hours = 0
            ws.cell(row=2 + i, column=9, value=hours)

            i += 1

        elif time_in == 18:
            ti = time(time_in).strftime("%H:%M")
            ci = clock_in

            if ci > ti:
                if ci[-2:] <= "05":
                    tti = 24.0 - float(ci[0:2])
                elif ci[-2:] <= "15":
                    tti = 24.0 - (float(ci[0:2]) + 0.25)
                elif ci[-2:] <= "30":
                    tti = 24.0 - (float(ci[0:2]) + 0.50)
                elif ci[-2:] <= "45":
                    tti = 24.0 - (float(ci[0:2]) + 0.75)
                elif ci[-2:] >= "45":
                    tti = 24.0 - (float(ci[0:2]) + 1.0)
            else:
                tti = 24.0 - float(time_in)

            hours = tti
            ws.cell(row=2 + i, column=9, value=hours)

            i += 1

        elif clock_in == None:
            # Calculate time out
            to = time(time_out).strftime("%H:%M")
            co = clock_out

            if co < to:
                if co[-2:] <= "05":
                    tto = float(co[0:2]) + 0.00
                elif co[-2:] <= "15":
                    tto = float(co[0:2]) + 0.25
                elif co[-2:] <= "30":
                    tto = float(co[0:2]) + 0.50
                elif co[-2:] <= "45":
                    tto = float(co[0:2]) + 0.75
                elif co[-2:] >= "45":
                    tto = (float(co[0:2]) + 1) + 0.00
            else:
                tto = int(time_out)

            hours = float(tto)
            ws.cell(row=2 + i, column=9, value=hours)

            i += 1

        else:
            ti = time(time_in).strftime("%H:%M")
            ci = clock_in

            to = time(time_out).strftime("%H:%M")
            co = clock_out

            # Calculate time in
            if ci > ti:
                if ci[-2:] <= "05":
                    tti = float(ci[0:2])
                elif ci[-2:] <= "15":
                    tti = float(ci[0:2]) + 0.25
                elif ci[-2:] <= "30":
                    tti = float(ci[0:2]) + 0.50
                elif ci[-2:] <= "45":
                    tti = float(ci[0:2]) + 0.75
                elif ci[-2:] >= "45":
                    tti = (float(ci[0:2]) + 1) + 0.00
            else:
                tti = float(time_in)

            # Calculate time out
            if co < to:
                if co[-2:] <= "05":
                    tto = float(co[0:2]) + 0.00
                elif co[-2:] <= "15":
                    tto = float(co[0:2]) + 0.25
                elif co[-2:] <= "30":
                    tto = float(co[0:2]) + 0.50
                elif co[-2:] <= "45":
                    tto = float(co[0:2]) + 0.75
                elif co[-2:] >= "45":
                    tto = (float(co[0:2]) + 1) + 0.00
            else:
                tto = int(time_out)

            # calculate hours worked
            hours = float(tto) - float(tti)
            ws.cell(row=2 + i, column=9, value=hours)

            i += 1

    wb.save("Wage Times.xlsx")
    wb.close()

def cas_public_weekone():
    wb = load_workbook("Wage Times.xlsx")
    ws = wb['Cashier Week One']

    i = 0

    for x in range(ws.max_row):
        date = ws.cell(row=2 + i, column=4).value
        hours = ws.cell(row=2 + i, column=9).value

        if date in public_holidays:
            ws.cell(row=2 + i, column=11, value=hours)
            ws.cell(row=2 + i, column=9, value='')
        
        i += 1

    wb.save("Wage Times.xlsx")
    wb.close()

def bak_cas_work():
    wb = load_workbook("Wage Times.xlsx")
    ws = wb['Cashier Week One']

    i = 0

    for x in range(ws.max_row):
        name = ws.cell(row=2 + i, column=1).value
        date = ws.cell(row=2 + i, column=4).value
        hours = ws.cell(row=2 + i, column=9).value

        for x in bc_working:
            if name == x[0] and date == x[1]:
                ws.cell(row=2 + i, column=13, value=hours)
                ws.cell(row=2 + i, column=9, value='')
        
        i += 1

    wb.save("Wage Times.xlsx")
    wb.close()

def cas_total_wo_hours():
    # Calculate total hours for week add to excel
    wb = load_workbook("Wage Times.xlsx")
    ws = wb['Cashier Week One']

    i = 0
    total = 0
    total_s = 0
    total_p = 0
    total_nc = 0
    total_bc = 0

    for x in range(ws.max_row):
        name = ws.cell(row=2 + i, column=1).value
        n = ws.cell(row=2 + i - 1, column=1).value
        badge = ws.cell(row=2 + i - 1, column=2).value
        day = ws.cell(row=2 + i, column=3).value
        hours = ws.cell(row=2 + i, column=9).value
        hours_s = ws.cell(row=2 + i, column=10).value
        hours_p = ws.cell(row=2 + i, column=11).value
        nc = ws.cell(row=2 + i, column=12).value
        bc_hours = ws.cell(row=2 + i, column=13).value

        # Check if name is true
        if name:
            if nc != None:
                total_nc = 1
            elif bc_hours !=None:
                total_bc += bc_hours
            elif day == 'Sunday':
                total_s += hours_s
            elif hours_p != None:
                total_p += hours_p
            elif hours == None:
                total += 0
            else:
                total += hours
            
            i += 1
        
        elif "Total" in n:
            i += 1

        else:
            ws.cell(row=2 + i, column=1, value= n + " " + "Total")
            ws.cell(row=2 + i, column=2, value=badge)
            ws.cell(row=2 + i, column=9, value=total)
            ws.cell(row=2 + i, column=10, value=total_s)
            ws.cell(row=2 + i, column=11, value=total_p)
            ws.cell(row=2 + i, column=12, value=total_nc)
            ws.cell(row=2 + i, column=13, value=total_bc)
            
            total = 0            
            total_s = 0  
            total_p = 0  
            total_nc = 0  
            total_bc = 0  

            i += 1

    wb.save("Wage Times.xlsx")
    wb.close()

# ==============================================================================
# CALCULATE CLOCK IN AND CLOCK OUT TIMES WEEK TWO
# ==============================================================================

def cas_times_weektwo():
    wb = load_workbook("Wage Times.xlsx")
    ws = wb['Cashier Week Two']

    i = 0
    for x in range(ws.max_row + 1):
        tti = 0
        tto = 0

        name = ws.cell(row=2 + i, column=1).value

        day = ws.cell(row=2 + i, column=3).value

        time_in = ws.cell(row=2 + i, column=5).value
        clock_in = ws.cell(row=2 + i, column=7).value

        time_out = ws.cell(row=2 + i, column=6).value
        clock_out = ws.cell(row=2 + i, column=8).value

        if name == None:
            i += 1

        elif time_in > 0 and clock_in == None or time_out > 0 and clock_out == None:
            hours = 'No Clock'
            ws.cell(row=2 + i, column=12, value=hours)

            i += 1

        elif day == "Sunday":
            if time_in == 18:
                ti = time(time_in).strftime("%H:%M")
                ci = clock_in

                if ci > ti:
                    if ci[-2:] <= "05":
                        tti = 24.0 - float(ci[0:2])
                    elif ci[-2:] <= "15":
                        tti = 24.0 - (float(ci[0:2]) + 0.25)
                    elif ci[-2:] <= "30":
                        tti = 24.0 - (float(ci[0:2]) + 0.50)
                    elif ci[-2:] <= "45":
                        tti = 24.0 - (float(ci[0:2]) + 0.75)
                    elif ci[-2:] >= "45":
                        tti = 24.0 - (float(ci[0:2]) + 1.0)
                else:
                    tti = 24.0 - float(time_in)

                hours = tti
                ws.cell(row=2 + i, column=10, value=hours)

            elif clock_in == None and clock_out == None:
                hours = 0
                ws.cell(row=2 + i, column=10, value=hours)

            elif clock_in == None:
                # Calculate time out
                to = time(time_out).strftime("%H:%M")
                co = clock_out

                if co < to:
                    if co[-2:] <= "05":
                        tto = float(co[0:2]) + 0.00
                    elif co[-2:] <= "15":
                        tto = float(co[0:2]) + 0.25
                    elif co[-2:] <= "30":
                        tto = float(co[0:2]) + 0.50
                    elif co[-2:] <= "45":
                        tto = float(co[0:2]) + 0.75
                    elif co[-2:] >= "45":
                        tto = (float(co[0:2]) + 1) + 0.00
                else:
                    tto = int(time_out)

                hours = float(tto)
                ws.cell(row=2 + i, column=10, value=hours)

            else:
                ti = time(time_in).strftime("%H:%M")
                ci = clock_in

                to = time(time_out).strftime("%H:%M")
                co = clock_out

                # Calculate time in
                if ci > ti:
                    if ci[-2:] <= "05":
                        tti = float(ci[0:2])
                    elif ci[-2:] <= "15":
                        tti = float(ci[0:2]) + 0.25
                    elif ci[-2:] <= "30":
                        tti = float(ci[0:2]) + 0.50
                    elif ci[-2:] <= "45":
                        tti = float(ci[0:2]) + 0.75
                    elif ci[-2:] >= "45":
                        tti = (float(ci[0:2]) + 1) + 0.00
                else:
                    tti = float(time_in)

                # Calculate time out
                if co < to:
                    if co[-2:] <= "05":
                        tto = float(co[0:2]) + 0.00
                    elif co[-2:] <= "15":
                        tto = float(co[0:2]) + 0.25
                    elif co[-2:] <= "30":
                        tto = float(co[0:2]) + 0.50
                    elif co[-2:] <= "45":
                        tto = float(co[0:2]) + 0.75
                    elif co[-2:] >= "45":
                        tto = (float(co[0:2]) + 1) + 0.00
                else:
                    tto = int(time_out)

                # calculate hours worked
                hours = float(tto) - float(tti)
                ws.cell(row=2 + i, column=10, value=hours)
            
            i += 1

        elif clock_in == None and clock_out == None:
            hours = 0
            ws.cell(row=2 + i, column=9, value=hours)

            i += 1

        elif time_in == 18:
            ti = time(time_in).strftime("%H:%M")
            ci = clock_in

            if ci > ti:
                if ci[-2:] <= "05":
                    tti = 24.0 - float(ci[0:2])
                elif ci[-2:] <= "15":
                    tti = 24.0 - (float(ci[0:2]) + 0.25)
                elif ci[-2:] <= "30":
                    tti = 24.0 - (float(ci[0:2]) + 0.50)
                elif ci[-2:] <= "45":
                    tti = 24.0 - (float(ci[0:2]) + 0.75)
                elif ci[-2:] >= "45":
                    tti = 24.0 - (float(ci[0:2]) + 1.0)
            else:
                tti = 24.0 - float(time_in)

            hours = tti
            ws.cell(row=2 + i, column=9, value=hours)

            i += 1

        elif clock_in == None:
            # Calculate time out
            to = time(time_out).strftime("%H:%M")
            co = clock_out

            if co < to:
                if co[-2:] <= "05":
                    tto = float(co[0:2]) + 0.00
                elif co[-2:] <= "15":
                    tto = float(co[0:2]) + 0.25
                elif co[-2:] <= "30":
                    tto = float(co[0:2]) + 0.50
                elif co[-2:] <= "45":
                    tto = float(co[0:2]) + 0.75
                elif co[-2:] >= "45":
                    tto = (float(co[0:2]) + 1) + 0.00
            else:
                tto = int(time_out)

            hours = float(tto)
            ws.cell(row=2 + i, column=9, value=hours)

            i += 1

        else:
            ti = time(time_in).strftime("%H:%M")
            ci = clock_in

            to = time(time_out).strftime("%H:%M")
            co = clock_out

            # Calculate time in
            if ci > ti:
                if ci[-2:] <= "05":
                    tti = float(ci[0:2])
                elif ci[-2:] <= "15":
                    tti = float(ci[0:2]) + 0.25
                elif ci[-2:] <= "30":
                    tti = float(ci[0:2]) + 0.50
                elif ci[-2:] <= "45":
                    tti = float(ci[0:2]) + 0.75
                elif ci[-2:] >= "45":
                    tti = (float(ci[0:2]) + 1) + 0.00
            else:
                tti = float(time_in)

            # Calculate time out
            if co < to:
                if co[-2:] <= "05":
                    tto = float(co[0:2]) + 0.00
                elif co[-2:] <= "15":
                    tto = float(co[0:2]) + 0.25
                elif co[-2:] <= "30":
                    tto = float(co[0:2]) + 0.50
                elif co[-2:] <= "45":
                    tto = float(co[0:2]) + 0.75
                elif co[-2:] >= "45":
                    tto = (float(co[0:2]) + 1) + 0.00
            else:
                tto = int(time_out)

            # calculate hours worked
            hours = float(tto) - float(tti)
            ws.cell(row=2 + i, column=9, value=hours)

            i += 1

    wb.save("Wage Times.xlsx")
    wb.close()

def cas_public_weektwo():
    wb = load_workbook("Wage Times.xlsx")
    ws = wb['Cashier Week Two']

    i = 0

    for x in range(ws.max_row):
        date = ws.cell(row=2 + i, column=4).value
        hours = ws.cell(row=2 + i, column=9).value

        if date in public_holidays:
            ws.cell(row=2 + i, column=11, value=hours)
            ws.cell(row=2 + i, column=9, value='')
        
        i += 1

    wb.save("Wage Times.xlsx")
    wb.close()

def bak_cas_work_wt():
    wb = load_workbook("Wage Times.xlsx")
    ws = wb['Cashier Week Two']

    i = 0

    for x in range(ws.max_row):
        name = ws.cell(row=2 + i, column=1).value
        date = ws.cell(row=2 + i, column=4).value
        hours = ws.cell(row=2 + i, column=9).value

        for x in bc_working:
            if name == x[0] and date == x[1]:
                ws.cell(row=2 + i, column=13, value=hours)
                ws.cell(row=2 + i, column=9, value='')
        
        i += 1

    wb.save("Wage Times.xlsx")
    wb.close()

def cas_total_wt_hours():
    # Calculate total hours for week add to excel
    wb = load_workbook("Wage Times.xlsx")
    ws = wb['Cashier Week Two']

    i = 0
    total = 0
    total_s = 0
    total_p = 0
    total_nc = 0
    total_bc = 0

    # print(ws.max_row)
    for x in range(ws.max_row):
        name = ws.cell(row=2 + i, column=1).value
        n = ws.cell(row=2 + i - 1, column=1).value
        badge = ws.cell(row=2 + i - 1, column=2).value
        day = ws.cell(row=2 + i, column=3).value
        nc = ws.cell(row=2 + i, column=12).value
        hours = ws.cell(row=2 + i, column=9).value
        hours_s = ws.cell(row=2 + i, column=10).value
        hours_p = ws.cell(row=2 + i, column=11).value
        bc_hours = ws.cell(row=2 + i, column=13).value
        
        # Check if name is true
        if name:
            if nc != None:
                total_nc = 1
            elif bc_hours != None:
                total_bc += bc_hours
            elif day == 'Sunday':
                total_s += hours_s
            elif hours_p != None:
                total_p += hours_p
            elif hours == None:
                total += 0
            else:
                total += hours
            
            i += 1
        
        elif "Total" in n:
            i += 1

        else:
            ws.cell(row=2 + i, column=1, value= n + " " + "Total")
            ws.cell(row=2 + i, column=2, value=badge)
            ws.cell(row=2 + i, column=9, value=total)
            ws.cell(row=2 + i, column=10, value=total_s)
            ws.cell(row=2 + i, column=11, value=total_p)
            ws.cell(row=2 + i, column=12, value=total_nc)
            ws.cell(row=2 + i, column=13, value=total_bc)
            
            total = 0            
            total_s = 0  
            total_p = 0  
            total_nc = 0  
            total_bc = 0  

            i += 1

    wb.save("Wage Times.xlsx")
    wb.close()

# ==============================================================================
# CALCULATE TOTAL HOURS WEEK ONE AND TWO
# ==============================================================================

def cas_total_wo_db():
    # Add Total week one to database
    wb = load_workbook("Wage Times.xlsx")
    ws = wb['Cashier Week One']

    # Add week one totals to total table in database
    # Connect to database
    con = sqlite3.connect("wageTimes.db")
    c = con.cursor()
    # Create table
    c.execute(
            """CREATE TABLE IF NOT EXISTS cashierTotal (
                name TEXT,
                badge TEXT,
                normal TEXT,
                sunday TEXT,
                public TEXT,
                noClock TEXT,
                cashier TEST
                )""")

    # Add week one data to table
    query = """INSERT INTO cashierTotal (
                name,
                badge,
                normal,
                sunday,
                public,
                noClock,
                cashier
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)"""

    i = 0
    for x in range(ws.max_row):
        name = ws.cell(row=2 + i, column=1).value
        badge = ws.cell(row=2 + i, column=2).value
        normal = ws.cell(row=2 + i, column=9).value
        sunday = ws.cell(row=2 + i, column=10).value
        public = ws.cell(row=2 + i, column=11).value
        nc = ws.cell(row=2 + i, column=12).value
        bc = ws.cell(row=2 + i, column=13).value

        if name != None:
            if 'Total' in name:
                c.execute(query, (name, badge, normal, sunday, public, nc, bc))
        
        i += 1

    # Save and close database
    con.commit()
    con.close()

    # Close workbook
    wb.save("Wage Times.xlsx")
    wb.close()

def cas_total_wt_db():
    # Add total week two to data base
    wb = load_workbook("Wage Times.xlsx")
    ws = wb['Cashier Week Two']

    # Add week one totals to total table in database
    # Connect to database
    con = sqlite3.connect("wageTimes.db")
    c = con.cursor()

    # Add week two totals to table
    query = ("""UPDATE cashierTotal
            SET
                normal = normal + ?,
                sunday = sunday + ?,
                public = public + ?,
                noClock = noClock + ?,
                cashier = cashier + ?
            WHERE
                badge = ?
                """)

    i = 0
    for x in range(ws.max_row):
        name = ws.cell(row=2 + i, column=1).value
        badge = ws.cell(row=2 + i, column=2).value
        normal = ws.cell(row=2 + i, column=9).value
        sunday = ws.cell(row=2 + i, column=10).value
        public = ws.cell(row=2 + i, column=11).value
        nc = ws.cell(row=2 + i, column=12).value
        bc = ws.cell(row=2 + i, column=13).value

        if name != None:
            if 'Total' in name:
                c.execute(query, (normal, sunday, public, nc, bc, badge))
        
        i += 1

    # Save and close database
    con.commit()
    con.close()

    # Close workbook
    wb.save("Wage Times.xlsx")
    wb.close()

def cas_fortnight_total():
    # Get totals from database
    con = sqlite3.connect("wageTimes.db")
    c = con.cursor()

    c.execute("SELECT * FROM cashierTotal")
    records = c.fetchall()

    # Save and close database
    con.commit()
    con.close()

    # Write to excel
    wb = load_workbook("Wage Times.xlsx")
    wb.create_sheet('Cashier Total')
    ws = wb['Cashier Total']

    # Create total sheet and headings
    ws["A1"] = 'Name'
    ws["B1"] = 'Total Normal Hours'
    ws["C1"] = 'Total Sunday Hours'
    ws["D1"] = 'Total Public Holiday Hours'
    ws["E1"] = 'No Clock'
    ws["F1"] = 'Baker/Cashier Hours'

    i = 0
    for r in records:
        name = r[0].replace('Total','')
        normal = float(r[2])
        sunday = float(r[3])
        public = float(r[4])
        no_clock = float(r[5])
        bc = float(r[6])

        ws.cell(row=2 + i, column=1, value=name)
        ws.cell(row=2 + i, column=2, value=normal)
        ws.cell(row=2 + i, column=3, value=sunday)
        ws.cell(row=2 + i, column=4, value=public)
        ws.cell(row=2 + i, column=6, value=bc)
        if no_clock == 1 or no_clock == 2:
            ws.cell(row=2 + i, column=5, value='No Clock')

        i += 1

    wb.save("Wage Times.xlsx")
    wb.close()

