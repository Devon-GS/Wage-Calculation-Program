import os
import re
import pandas as pd
import database as db
from datetime import datetime, timedelta, time
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from config import (CREATE_EXCEL,CREATE_CARWASH_TIMES, DYNAMIC_FILE_LOC, WAGE_TIMES_FILE, PUBLIC_HOILIDAY_FILE, UNICLOX_FOLDER,
					ATT_ROSTER_FILE, CAS_ROSTER_FILE, BADGE_NUMBER_FILE, BAKER_CASHIER_FILE, CARWASH_FILE, COLUMN_WIDTHS_ATT, 
					COLUMN_WIDTHS_TOTALS ,COL_DIFF)

# --- Helper Functions ---
def clear_excel():
	if os.path.isfile(WAGE_TIMES_FILE):
		os.remove(WAGE_TIMES_FILE)

def clear_carwash_excel():
	if os.path.isfile(CARWASH_FILE):
		os.remove(CARWASH_FILE)

def load_excel():
	"""Opens the workbook and returns the object."""
	if not os.path.isfile(WAGE_TIMES_FILE):
		CREATE_EXCEL()

	wb = load_workbook(WAGE_TIMES_FILE) 
	return wb  

def save_workbook(wb):
	"""Saves the workbook to the disk."""
	wb.save(WAGE_TIMES_FILE)
	wb.close()

def get_badge_mapping():
	"""Creates a dictionary {Name: BadgeID} from the badges.xlsx file."""
	mapping = {}
	if os.path.exists(BADGE_NUMBER_FILE):
		# Load without headers
		df = pd.read_excel(BADGE_NUMBER_FILE, header=None)
		for index, row in df.iterrows():
			# row[0] is Name, row[1] is Badge
			mapping[str(row[0]).strip()] = str(row[1]).strip()
	return mapping

def get_cashier_dates():
	"""Gets dates of cashier shifts for employee that works cashier and baker shifts"""
	if os.path.exists(BAKER_CASHIER_FILE):
		wb = load_workbook(BAKER_CASHIER_FILE, data_only=True)
		ws = wb.active

		bc_working = []
		for row in ws.iter_rows(min_row=2, max_col=2, max_row=20, values_only=True):
			x = row
			if x[0] != None:
				name = x[0]
				cashier_date = x[1].strftime('%d/%m/%Y')
				bc = [name, cashier_date]
				bc_working.append(bc)

		wb.close()
		return bc_working

def split_roster_time(val):
	"""Replaces the old first() and second() regex functions."""
	if val in ["AF", " ", "0", 0, None, ""]:
		return 0.0, 0.0
	try:
		# Matches "08-17" or "18-06"
		times = re.findall(r"(\d+)", str(val))
		return int(times[0]), int(times[1])
	except:
		return 0.0, 0.0

def collect_public_holidays():
	"""Get public holidays and save to database"""
	holidays = []
	if os.path.exists(PUBLIC_HOILIDAY_FILE):
		wb = load_workbook(PUBLIC_HOILIDAY_FILE, data_only=True)
		ws = wb.active
		for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
			if row[0]: 
				holidays.append((row[0].strftime('%d/%m/%Y'),))
		wb.close()
	db.public_holidays_db(holidays)

def adjust_time(clock_hours, roster_h, day, date, holidays, is_in):
	"""
	1. Rounding logic - changes the dicimal to 15, 30 or 45
	2. m = Minutes and h = Hours 
	3. is_in = Clock in or out
	4. Check if Sunday or public holiday and gives no leeway
	"""

	# Check if manual clock for recaculation
	if type(clock_hours) != str:
		clock_hours = clock_hours.strftime("%H:%M")

	# Set flag
	if day == 'Sunday':
		flag = 'sun'
	elif date in holidays:
		flag = 'pub'
	else:
		flag = 'norm'

	if not clock_hours:
		return float(roster_h), flag
	
	# Split hours and minites
	h, m = map(int, clock_hours.split(':'))

	# Clock In Logic
	if is_in: 
		if h > roster_h or (h == roster_h and m > 0):
			# Special logic for Sunday: No 4-minute grace period
			if date in holidays or day == "Sunday":
				# Set flag
				if day == "Sunday":
					flag = 'sun'
				else:
					flag = 'pub'

				if m <= 15: 
					return h + 0.25, flag
				elif m <= 30: 
					return h + 0.50, flag
				elif m <= 45: 
					return h + 0.75, flag
				else: 
					return float(h + 1), flag
			 # Standard logic for all other days
			else:
				if m <= 4: 
					return float(h), flag # Gives employee 4 min to clock in
				elif m <= 15: 
					return h + 0.25, flag
				elif m <= 30: 
					return h + 0.50, flag
				elif m <= 45: 
					return h + 0.75, flag
				else: 
					return float(h + 1), flag
		return float(roster_h), flag
	# Clock Out Logic
	else: 
		if h < roster_h:
			if m <= 15: 
				return float(h), flag
			elif m <= 30: 
				return h + 0.25, flag
			elif m <= 45: 
				return h + 0.50, flag
			else: 
				return h + 0.75, flag
		return float(roster_h), flag


# --- Helper Functions End ---


# --- Step 1: Roster to Database ---
def roster_shift_to_db(role="Attendant", week="WeekOne"):
	"""
	1. Reads the Roster (Attendant or Cashier).
	2. Gets the Badge Mapping.
	3. Get roster shifts week one and two for attendants and cashiers.
	4. Writes names, badges, days, dates and shift to database.
	"""
	# Get Path and Badge Mapping
	file_path = ATT_ROSTER_FILE if role == "Attendant" else CAS_ROSTER_FILE
	badges = get_badge_mapping()
	
	# Load the Roster via Pandas
	try:
		if role == "Attendant" and week == "WeekOne":
			# Columns used
			cols = "B:I"

			# Dates slice
			drow = 0
			d_col_start = 1
			d_col_end = 8

			# Week times slice
			wrow = 2
			wrow_end = 17

		# Att week two
		elif role == "Attendant" and week == "WeekTwo":
			# Columns used
			cols = "B:I"
			
			# Dates slice
			drow = 28
			d_col_start = 1
			d_col_end = 8

			# Week times slice
			wrow = 30
			wrow_end = 45

		# Cashier week one 
		elif role == "Cashier" and week == "WeekOne":
			# Columns used
			cols = "B:I"
			
			# Dates slice
			drow = 3
			d_col_start = 1
			d_col_end = 8

			# Week times slice
			wrow = 5
			wrow_end = 11

			# Week times slice (bakers)
			wbrow = 31
			wbrow_end = 33
		
		# Cashier week two 
		elif role == "Cashier" and week == "WeekTwo":
			# Columns used
			cols = "B:I"
			
			# Dates slice
			drow = 34
			d_col_start = 1
			d_col_end = 8

			# Week times slice
			wrow = 14
			wrow_end = 20

			# Week times slice (bakers)
			wbrow = 36	
			wbrow_end = 38
		
		# Get times from excel
		df = pd.read_excel(file_path, header=None, usecols=cols, nrows=46)
		data = df.fillna(0)

		# Extract the dates  
		week_dates = data.iloc[drow, d_col_start : d_col_end]

		# Extract the employee schedule block
		if role == "Cashier":
			cashier_times = data.iloc[wrow : wrow_end].copy()
			
			# Add bakers 
			baker_times = data.iloc[wbrow : wbrow_end]

			# Combine them
			week_times = pd.concat([cashier_times, baker_times])
		else:
			week_times = data.iloc[wrow : wrow_end]		

		# Create an empty list to store the final tuples
		schedule_list = []

		# Iterate through every row in the week_times dataframe
		for index, row in week_times.iterrows():
			name = row[1]  # Column 0 (Excel column B) contains the employee names
			
			# Check if we have a valid name (skip empty rows filled with 0)
			if str(name) != 'nan' and name != 0:
				
				# Iterate over the column indices where we know the dates are (1 through 7)
				for col_idx in week_dates.index:
					shift = row[col_idx]
			
					# If the employee actually has a shift that day (not 0)
					if shift != 0 and str(shift) != 'nan':
						date_obj = week_dates[col_idx]

						# Convert the string/object to a reliable pandas datetime object
						dt_obj = pd.to_datetime(date_obj, dayfirst=True, errors='coerce')
						
						# Format the date and day (assuming date_obj is a datetime object)
						if pd.notna(dt_obj):
							day_name = dt_obj.strftime("%A").capitalize()  # e.g., "Monday"
							date_str = dt_obj.strftime("%d/%m/%Y")    # e.g., "03/03/2026"

							# Get badge
							badge_id = badges.get(name, "NOT FOUND")

							schedule_list.append((name, badge_id, day_name, date_str, shift, week))
	
		# Add shifts to database
		db.add_shifts(schedule_list, role, week)

	except Exception as e:
		print(f"Error initializing roster: {e}")

# --- Step 2: Clock Collection (Logic from att_clock_times.py) ---
def collect_clock_times():
	"""Reads last 5 files from Uniclox folder and saves to DB."""
	clock_times = []
	clock_files = [f for f in os.listdir(UNICLOX_FOLDER) if 'TL' in f and f[-7:-4] != '000']
	recent_files = clock_files[-5:]

	for filename in recent_files:
		with open(os.path.join(UNICLOX_FOLDER, filename), 'r') as f:
			for line in f:
				parts = line.strip().split(',')
				if len(parts) < 2: continue
				badge = parts[0]
				dt_obj = datetime.strptime(parts[1], '%Y-%m-%d %H:%M:%S')
				clock_times.append((badge, dt_obj.strftime("%d/%m/%Y"), dt_obj.strftime("%H:%M:%S")))

	db.add_clock_times(clock_times)

# --- Step 3: Write Roster Shifts to Excel (Logic from cas_clock_times.py) ---
def sync_shifts_to_excel(wb, sheet_name):
	"""
	1. Write roster data for name, badges, dates, shift times.
	2. Adds total headings under weekly shifts	
	"""
	# wb = load_workbook(WAGE_TIMES_FILE)
	ws = wb[sheet_name]

	# Get selected data from database
	if sheet_name == "Att Week One":
		data = db.get_shift_times_db('Attendant', 'WeekOne')
	elif sheet_name == "Att Week Two":
		data = db.get_shift_times_db('Attendant', 'WeekTwo')
	elif sheet_name == "Cashier Week One":
		data = db.get_shift_times_db('Cashier', 'WeekOne')
	else:
		data = db.get_shift_times_db('Cashier', 'WeekTwo')

	# Write shifts, badges, days, dates to Excel 
	current_row = 2
	prev_name = None
	prev_shift = 0

	for row in data:
		name = row[0]
		badge = row[1]
		day = row[2]
		date = row[3]
		shift = row[4]

		# Skip two lines between employees
		if prev_name == None:
			pass
		elif prev_name != name:
			ws.cell(row=current_row, column=1, value=f'{prev_name.upper()} Total')
			prev_shift = 0
			current_row += 2
		
		# Split shift times into start and end
		shift_times = split_roster_time(shift)
		shift_start = shift_times[0]
		shift_end = shift_times[1]	
				
		# Logic for night shift
		if shift_start == 0 and prev_shift >= 18:
			current_row -= 1
		
		elif shift_start >= 18:
			# Shift start
			ws.cell(row=current_row, column=1, value=name)
			ws.cell(row=current_row, column=2, value=badge)
			ws.cell(row=current_row, column=3, value=day)
			ws.cell(row=current_row, column=4, value=date)
			ws.cell(row=current_row, column=5, value=shift_start)
			ws.cell(row=current_row, column=6, value=0)

			# Get shift end and next day 
			# Convert the string into a datetime object (format: day/month/year)
			date_obj = datetime.strptime(date, "%d/%m/%Y")
		
			# Add one day
			next_date = date_obj + timedelta(days=1)
			new_date = next_date.strftime("%d/%m/%Y")

			# Get the day name (e.g., Thursday)
			day_name = next_date.strftime("%A")

			# Shift end
			ws.cell(row=current_row + 1, column=1, value=name)
			ws.cell(row=current_row + 1, column=2, value=badge)
			ws.cell(row=current_row + 1, column=3, value=day_name)
			ws.cell(row=current_row + 1, column=4, value=new_date)
			ws.cell(row=current_row + 1, column=5, value=0)
			ws.cell(row=current_row + 1, column=6, value=shift_end)

			current_row += 1

		else:
			ws.cell(row=current_row, column=1, value=name)
			ws.cell(row=current_row, column=2, value=badge)
			ws.cell(row=current_row, column=3, value=day)
			ws.cell(row=current_row, column=4, value=date)
			ws.cell(row=current_row, column=5, value=shift_start)
			ws.cell(row=current_row, column=6, value=shift_end)
		 
		current_row += 1
		prev_name = name
		prev_shift = shift_start

	ws.cell(row=current_row, column=1, value=f'{prev_name.upper()} Total')

# --- Step 3: Write Clocks to Excel ---
def sync_clocks_to_excel(wb, sheet_name):
	"""
	1. Matches clockings to the shift rows in the Excel sheet.
	2. ti = Time in | to = Time out [Actual clock times]
	3. Moves Sunday and Public holiday to right columns
	4. For cahsiers moves cashier/baker employee's cashiers times to right column
	"""

	clocks = db.get_clock_times()

	# wb = load_workbook(WAGE_TIMES_FILE)
	ws = wb[sheet_name]
	
	for i in range(2, ws.max_row + 1):
		badge = ws.cell(row=i, column=2).value
		date = ws.cell(row=i, column=4).value
		if not badge or not date: continue

		clocking_times = []

		for c_badge, c_date, c_time in clocks:
			if badge == c_badge and date == c_date:
				clocking_times.append(c_time)

		ti_roster = ws.cell(row=i, column=5).value
		to_roster = ws.cell(row=i, column=6).value

		# If shift is 'AF' ignore all clocking
		if ti_roster == 0.0 and to_roster == 0.0:
			continue

		if not clocking_times:
			continue

		# Handle night shift
		elif ti_roster == 18:
			t = time.fromisoformat(max(clocking_times))

			# Checks double night shift if only one clock
			if t.hour > 14:
				t = time.fromisoformat(max(clocking_times)).strftime('%H:%M')
				ws.cell(row=i, column=7, value=t)

		# Handle morning of night shift
		elif to_roster == 6 or to_roster == 7:
			t = time.fromisoformat(min(clocking_times))

			# Checks double night shift if only one clock
			if t.hour < 14:
				t = time.fromisoformat(min(clocking_times)).strftime('%H:%M')
				ws.cell(row=i, column=8, value=t)

		# Handle when employee clocks in/out only once 
		elif ti_roster > 0 and to_roster > 0 and len(clocking_times) == 1:
			# Logic for picking min/max based on shift
			# Single clocking: Determine if it's an IN or an OUT
			clock_h = int(clocking_times[0].split(':')[0])
			if abs(clock_h - (ti_roster or 0)) < abs(clock_h - (to_roster or 0)):
				t = time.fromisoformat(clocking_times[0]).strftime('%H:%M')
				ws.cell(row=i, column=7, value=t)
			else:
				t = time.fromisoformat(clocking_times[0]).strftime('%H:%M')
				ws.cell(row=i, column=8, value=t)
		
		# Handle when employee clocks in/out multiple times but does not clock in/out
		elif ti_roster > 0 and to_roster > 0 and len(clocking_times) >= 2:
			# Get clock times
			clock_min = (min(clocking_times).split(':')[0])
			clock_max = (max(clocking_times).split(':')[0])
			
			# Get roster shifts 
			roster_min = ws.cell(row=i, column=5).value
			roster_max = ws.cell(row=i, column=6).value

			# Check if the two clock times match
			if clock_min == clock_max:
				# Find if  clocks where start or end shift
				low = int(roster_min) - int(clock_min)
				high = int(roster_max) - int(clock_min)

				# Start shift
				if abs(low) < abs(high):
					t = time.fromisoformat(clocking_times[0]).strftime('%H:%M')
					ws.cell(row=i, column=7, value=t)
				else:
					t = time.fromisoformat(clocking_times[0]).strftime('%H:%M')
					ws.cell(row=i, column=8, value=t)
			# Handle normal shift
			else:
				ti = time.fromisoformat(min(clocking_times)).strftime('%H:%M')
				ws.cell(row=i, column=7, value=ti)

				to = time.fromisoformat(max(clocking_times)).strftime('%H:%M')
				ws.cell(row=i, column=8, value=to)

# --- Step 4: Calculate Hours (Logic from att_cal_hours.py) ---
def calculate_hours(wb, sheet_name):
	"""
	1. Calculates shift vs clocking hours
	2. Calculate total normal, sunday, public hours
	"""
	# wb = load_workbook(WAGE_TIMES_FILE)
	ws = wb[sheet_name]

	# Get baker's cashier hours
	bc = get_cashier_dates()

	# Get public holidays
	holidays = db.get_public_holidays()

	# -- Caculate Shift vs Clocking Times To Get Hours Worked ---
	for i in range(2, ws.max_row + 1):
		name = ws.cell(row=i, column=1).value
		if not name or 'Total' in name: 
			continue

		day = ws.cell(row=i, column=3).value
		date = ws.cell(row=i, column=4).value
		ti = ws.cell(row=i, column=5).value  # Roster In
		to = ws.cell(row=i, column=6).value  # Roster Out
		ci = ws.cell(row=i, column=7).value  # Clock In (str HH:MM)
		co = ws.cell(row=i, column=8).value  # Clock Out (str HH:MM)

		# Checks to see if an employee did not clock
		if (ti and ti > 0 and not ci) or (to and to > 0 and not co):
			ws.cell(row=i, column=12, value="No Clock")
			continue

		# Rounding Logic and spliting time and flag
		calc_ti = adjust_time(ci, ti, day, date, holidays, True) if ci else 0
		calc_to = adjust_time(co, to, day, date, holidays, False) if co else 0
		
		# Set flag for shifts where employee is off
		if calc_ti == 0:
			if day == 'Sunday':
				calc_ti = (0, 'sun')
			elif date in holidays:
				calc_ti = (0, 'pub')
			else:
				calc_ti = (0, 'norm')

		if calc_to == 0:
			if day == 'Sunday':
				calc_to = (0, 'sun')
			elif date in holidays:
				calc_to = (0, 'pub')
			else:
				calc_to = (0, 'norm')

		calc_ti_t = calc_ti[0]		# Set time
		calc_ti_f = calc_ti[1]		# Set flag

		calc_to_t = calc_to[0]		# Set time
		calc_to_f = calc_to[1] 		# Set flag

		# Night Shift Logic
		if ti == 18:
			hours = 24.0 - calc_ti_t
		elif ti == 0 and to > 0:
			hours = calc_to_t
		else:
			hours = calc_to_t - calc_ti_t

		# Assign columns
		# Check if employee is a baker and a cashier
		if [name, date] in bc:
			if calc_ti_f == 'pub' or calc_to_f == 'pub':
				ws.cell(row=i, column=9, value='')
				ws.cell(row=i, column=15, value=hours)
			elif calc_ti_f == 'sun' or calc_to_f == 'sun':
				ws.cell(row=i, column=9, value='')
				ws.cell(row=i, column=14, value=hours)
			else:
				ws.cell(row=i, column=9, value='')
				ws.cell(row=i, column=13, value=hours)
		# All other employees
		elif calc_ti_f == 'pub' or calc_to_f == 'pub':
			ws.cell(row=i, column=9, value='')
			ws.cell(row=i, column=11, value=hours)
		elif calc_ti_f == 'sun' or calc_to_f == 'sun':
			ws.cell(row=i, column=9, value='') 
			ws.cell(row=i, column=10, value=hours)
		else: 
			ws.cell(row=i, column=9, value=hours)

# --- Step 5: Total Hours Worked ---
def cal_total_hours(wb, role="Attendant"):
	# Check what role is being calculated
	if role == "Attendant":
		sheets = ['Att Week One', 'Att Week Two']
		total_sheet = "Att Total" 
	else:	
		sheets = ['Cashier Week One', 'Cashier Week Two']
		total_sheet = "Cashier Total" 

	# Initilize totals dic
	totals = {}	

	# Loop through sheets an calulate totals
	for sheet in sheets:
		ws = wb[sheet]

		w_totals = {}	
	
		# Iterate through rows (start at row 2 to skip headers)
		# Using ws.max_row + 1 to ensure the last person's total is written
		for row in range(2, ws.max_row + 2):
			name = ws.cell(row=row, column=1).value
			day = ws.cell(row=row, column=3).value

			# Determine if this is a "Total" row or an empty break row
			is_total_row = name and "Total" in str(name)
			# is_empty_row = name is None

			# If it's a normal day row, accumulate hours - Adds daily hours
			if name and not is_total_row:
				# Create name key in dic
				if role == 'Attendant':
					w_totals.setdefault(name, {'badge':0, 'std': 0, 'sun': 0, 'pub': 0, 'nc': 0})
				else:
					w_totals.setdefault(name, {'badge':0, 'std': 0, 'sun': 0, 'pub': 0, 'nc': 0, 'cstd':0, 'csun':0, 'cpub':0})

				# Accumulate values
				nc = ws.cell(row=row, column=12).value
				if nc is not None:
					w_totals[name]['nc'] = 1
				elif day == 'Sunday' and ws.cell(row=row, column=10).value is not None:
					w_totals[name]['sun'] += (ws.cell(row=row, column=10).value or 0)
				elif ws.cell(row=row, column=11).value is not None:
					w_totals[name]['pub'] += ws.cell(row=row, column=11).value
				elif role != 'Attendant':
					w_totals[name]['std'] += (ws.cell(row=row, column=9).value or 0)
					w_totals[name]['cstd'] += ws.cell(row=row, column=13).value or 0
					w_totals[name]['csun'] += ws.cell(row=row, column=14).value or 0
					w_totals[name]['cpub'] += ws.cell(row=row, column=15).value or 0
				else:
					w_totals[name]['std'] += (ws.cell(row=row, column=9).value or 0)

			# Write weekly total hours at end of week and adds the two weeks total hours 
			elif name and is_total_row:
				# Get name without 'Total'
				name_total = ws.cell(row=row - 1, column=1).value
				# Get badge of empoyee
				badge = ws.cell(row=row - 1, column=2).value 

				# Add weekly to total coloumn in excel
				ws.cell(row=row, column=9, value=w_totals[name_total]['std'])
				ws.cell(row=row, column=10, value=w_totals[name_total]['sun'])
				ws.cell(row=row, column=11, value=w_totals[name_total]['pub'])
				ws.cell(row=row, column=12, value=w_totals[name_total]['nc'])
				if role != 'Attendant':
					ws.cell(row=row, column=13, value=w_totals[name_total]['cstd'])
					ws.cell(row=row, column=14, value=w_totals[name_total]['csun'])
					ws.cell(row=row, column=15, value=w_totals[name_total]['cpub'])

				# Add two weeks to totals dic
				if role == "Attendant":
					totals.setdefault(name_total, {'badge':0, 'std': 0, 'sun': 0, 'pub': 0, 'nc': 0})
				else:
					totals.setdefault(name_total, {'badge':0, 'std': 0, 'sun': 0, 'pub': 0, 'nc': 0, 'cstd':0, 'csun':0,'cpub':0})
				
				totals[name_total]['badge'] = badge 

				totals[name_total]['std'] += w_totals[name_total]['std']
				totals[name_total]['sun'] += w_totals[name_total]['sun']
				totals[name_total]['pub'] += w_totals[name_total]['pub']
				totals[name_total]['nc'] += w_totals[name_total]['nc']
				if role != 'Attendant':
					totals[name_total]['cstd'] += w_totals[name_total]['cstd']
					totals[name_total]['csun'] += w_totals[name_total]['csun']
					totals[name_total]['cpub'] += w_totals[name_total]['cpub']
	
	# Send total to database
	db.add_total_hours_db(totals, role)

	# Sync totals to excel sheets 
	ws = wb[total_sheet]

	current_row = 2

	for name, hours in totals.items():
		ws.cell(row=current_row, column=1, value=name)
		ws.cell(row=current_row, column=2, value=hours['std'])
		ws.cell(row=current_row, column=3, value=hours['sun'])
		ws.cell(row=current_row, column=4, value=hours['pub'])
		if hours['nc'] == 1:
			ws.cell(row=current_row, column=5, value="No Clock")
		else:
			ws.cell(row=current_row, column=5, value="")
		if role != 'Attendant':
			ws.cell(row=current_row, column=6, value=hours['cstd'])
			ws.cell(row=current_row, column=7, value=hours['csun'])
			ws.cell(row=current_row, column=8, value=hours['cpub'])

		current_row += 1

# --- Step 4: Formating Excel (Logic from att_cal_hours.py) ---
def format_excel(wb):
	# Get column configs
	cols_att = COLUMN_WIDTHS_ATT
	cols_tot = COLUMN_WIDTHS_TOTALS
	col_diff = COL_DIFF

	# Sheet names
	weekly_sheets = ['Att Week One', 'Att Week Two', 'Cashier Week One', 'Cashier Week Two']
	total_sheets = ['Att Total', 'Cashier Total']

	# Apply formats to sheets
	for sheet_name in weekly_sheets + total_sheets:
		if sheet_name not in wb.sheetnames: 
			continue
		
		ws = wb[sheet_name]

		if sheet_name in weekly_sheets:
			# Apply Column Widths
			for col, size in cols_att.items():
				ws.column_dimensions[col].width = size + col_diff
			
			# Style 'Total' rows
			style_cols = [1, 2, 9, 10, 11, 12, 13, 14, 15] if 'Cashier' in sheet_name else [1, 2, 9, 10, 11, 12]

			for row in range(2, ws.max_row + 1):
				if ws.cell(row=row, column=1).value and 'Total' in str(ws.cell(row=row, column=1).value):
					for c in style_cols:
						ws.cell(row=row, column=c).style = "total_style"
		
		else: # Logic for Total sheets
			# Apply Column Widths
			for col, size in cols_tot.items():
				ws.column_dimensions[col].width = size + col_diff
			
			# Center Align columns B through F
			for row in range(2, ws.max_row + 1):
				for col_idx in range(2, 9):
					ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal='center')

# --- Step 5: Carwash (Logic from carwash_db.py) ---
def carwash_work_hours():
	# Remove old and recreate new carwash times
	clear_carwash_excel()
	CREATE_CARWASH_TIMES()
	
	# Read data from carwash hours 
	df = pd.read_excel(DYNAMIC_FILE_LOC('Carwash'), header=None, usecols='A:J', nrows=20)
	
	# 1. FIX THE HEADERS
	# Set the column names using the first row (index 0)
	df.columns = df.iloc[0]

	# 2. EXTRACT THE BADGE ROW 
	# This is the second row in the excel sheet (index 1)
	badge_row = df.iloc[1]

	# 3. EXTRACT THE DATA ROWS
	# These are the dates and hours (everything from index 2 onwards)
	data_rows = df.iloc[2:].copy()

	# 4. DYNAMICALLY IDENTIFY NAME COLUMNS
	# Take all columns after 'TOTAL' (index 2 onwards) 
	# and make sure they are valid strings (not 0.0 or empty)
	name_cols = [col for col in df.columns[2:] if isinstance(col, str) and col.strip() != '' and col != 0.0]

	# 5. MELT THE DATAFRAME
	df_long = data_rows.melt(
		id_vars=['DATE'], 
		value_vars=name_cols,
		var_name='Name', 
		value_name='Hours'
	)

	# 6. ADD THE BADGES
	# Create a dictionary from the badge row to map Name -> Badge
	badge_map = badge_row.to_dict()
	df_long['Badge'] = df_long['Name'].map(badge_map)

	# 7. CLEANUP
	# Remove extra spaces from names and convert Badge to integer
	df_long['Name'] = df_long['Name'].str.strip()
	df_long['Badge'] = pd.to_numeric(df_long['Badge']).astype(int)

	# Final Order: Name, Badge, Date, Hours
	result = df_long[['Name', 'Badge', 'DATE', 'Hours']]

	# 8. RETURN AS DICTIONARY
	# 'records' -> [{column: value}, {column: value}]
	# return result.to_dict(orient='records')
	result_dic = result.to_dict(orient='records')

	carwash_total_hours = {}

	for x in result_dic:
		name = x['Name']
		badge = x['Badge']
		date = x['DATE']
		hours = x['Hours']

		# If badge not in dic create 
		if badge not in carwash_total_hours:
			carwash_total_hours[badge] = {}
			carwash_total_hours[badge]['date'] = {}
			carwash_total_hours[badge]['sun'] = 0
			carwash_total_hours[badge]['norm'] = 0

		# Add name and date and total norm and sun hours
		carwash_total_hours[badge]['name'] = name 
		carwash_total_hours[badge]['date'][date] = hours

		if date.strftime('%A') == 'Sunday':
			carwash_total_hours[badge]['sun'] += hours
		else: 
			carwash_total_hours[badge]['norm'] += hours 


	# -- WRITE WEEK DAY TIMES TO EXCEL (CARWASH_FILE) --- 
	# Constants 
	WEEK1_COL = 1      # Column A
	WEEK2_COL = 7      # Column G
	SUMMARY_COL = 13   # Column M

	wb = load_workbook(CARWASH_FILE)
	ws = wb['Times']	

	# Start rows
	current_base_row = 2 
	total_row = 2 
	extra_row = 13 

	for badge, info in carwash_total_hours.items():
		# Sort the dates chronologically to guarantee correct week grouping
		sorted_dates = sorted(info['date'].items())
		
		for i, (date_obj, hours) in enumerate(sorted_dates):
			
			week_num = i // 7 
			day_offset = i % 7
			
			target_row = current_base_row + day_offset
			col_start = WEEK2_COL if week_num else WEEK1_COL

			# Write daily info
			ws.cell(row=target_row, column=col_start, value=info['name'])
			ws.cell(row=target_row, column=col_start + 1, value=badge)
			ws.cell(row=target_row, column=col_start + 2, value=date_obj.strftime('%A'))
			
			# BEST PRACTICE: Write actual date objects and apply Excel number formatting
			date_cell = ws.cell(row=target_row, column=col_start + 3, value=date_obj)
			date_cell.number_format = 'DD/MM/YYYY'
			
			ws.cell(row=target_row, column=col_start + 4, value=hours)

		# Move to the next employee block
		current_base_row += 8

		# Write Total Hours block
		ws.cell(row=total_row, column=SUMMARY_COL, value=info['name'])
		ws.cell(row=total_row, column=SUMMARY_COL + 1, value=badge)
		ws.cell(row=total_row, column=SUMMARY_COL + 2, value=info['norm'])
		ws.cell(row=total_row, column=SUMMARY_COL + 3, value=info['sun'])

		# Write Extra Details block
		ws.cell(row=extra_row, column=SUMMARY_COL, value=info['name'])
		ws.cell(row=extra_row, column=SUMMARY_COL + 1, value=badge)
		
		total_row += 1
		extra_row += 1

		wb.save(CARWASH_FILE)
		wb.close()

def carwash_times():
	# Load the workbook
	wb = load_workbook(CARWASH_FILE, data_only=True)
	
	try:
		ws = wb['Times']
		data = {}

		# 1. Get normal and sunday hours
		for name, badge, norm, sun in ws.iter_rows(min_row=2, max_row=9, min_col=13, max_col=16, values_only=True):
			if name:
				# 2. Create dictionary 
				data[badge] = {
					'name': name,
					'n_hours': norm,
					's_hours': sun,
					'amount': 0  # Default value 
				}

		# 3. Add extra time
		for _, ebadge, _, amount in ws.iter_rows(min_row=13, max_row=20, min_col=13, max_col=16, values_only=True):
			if ebadge in data:
				data[ebadge]['amount'] = amount

		# Add carwash times to database
		db.carwash_db(data)
	finally:
		wb.close()