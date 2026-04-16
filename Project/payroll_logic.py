import os
import re
import xlwings as xw
import database as db
import pandas as pd
from openpyxl.utils import get_column_letter
from CTkMessagebox import CTkMessagebox
from openpyxl import load_workbook
from config import TAX_RATES_FILE, TAX_RESULTS


def run_payroll(PAYROLL_FILE):
	try:
		wb = load_workbook(PAYROLL_FILE)
		ws = wb['Wages']

		# Get all total hours
		total_records =  db.get_total_hours()

		# 1. Pre-process records into a dictionary
		records_dict = {str(r[1]): r for r in total_records}

		# Iterate over payroll columns starting from column C
		columns = ws.iter_cols(min_row=1, min_col=3)

		for col in columns:
			# col[1] is Row 2 in Excel (0-indexed tuple)
			badge_val = col[1].value 
			
			if badge_val is None:
				continue

			# Convert to string and clean it up
			badge_str = str(badge_val).strip().lower()

			# 2. Figure out the badge number and the match type
			if badge_str.endswith('c'):
				badge_key = badge_str[:-1]
				match_type = 'c'
			elif badge_str.endswith('b'):
				badge_key = badge_str[:-1]
				match_type = 'b'
			else:
				# Strips the .0 if it exists so it matches r[1]
				try:
					badge_key = str(int(float(badge_val)))
				except ValueError:
					badge_key = badge_str
				match_type = 'exact'

			# 3. Look up the record
			r = records_dict.get(badge_key)

			# Skip to the next column if no matching record is found
			if not r:
				continue  

			# Asign to variables
			norm = float(r[2])
			sun = float(r[3])
			pub = float(r[4])

			if int(badge_key) > 1000:
				extra_pay = int(r[5])

			# 4. Apply the payment logic
			# Baker's cashier hours
			if match_type == 'c':
				cnorm = float(r[6])
				csun = float(r[7])
				cpub = float(r[8])

				# Handle 'Null' to prevent ValueError crashes
				col[2].value  = cnorm
				col[11].value = csun
				col[14].value = cpub

			# Baker's baker hours
			elif match_type == 'b':
				col[2].value  = norm
				col[11].value = sun
				col[14].value = pub

			# Carwash hours 
			elif match_type == 'exact' and int(badge_key) > 1000:
				col[2].value  = norm
				col[11].value = sun
				col[14].value = pub
				col[20].value = extra_pay

			# Every one else
			else:
				col[2].value  = norm
				col[11].value = sun
				col[14].value = pub	
		
		wb.save(PAYROLL_FILE)
		wb.close()
	except Exception as error:
		CTkMessagebox(title="Error", message=str(error), icon="cancel")


# =========================================================================================
#  CALCULATE TAX
# =========================================================================================

# --- HELPER FUNCTIONS ---

def clean_currency(x):
	"""Cleans commas, spaces, and currency symbols, returning an integer."""
	if isinstance(x, str):
		x = x.replace(',', '').replace('R', '').replace(' ', '')
	try:
		return int(float(x))
	except (ValueError, TypeError):
		return 0

def recalculate_excel_formulas(filepath):
	"""Opens and saves an Excel file in the background to force formula recalculation."""
	with xw.App(visible=False) as app:
		book = app.books.open(filepath)
		book.save()
		book.close()

def get_tax_amount(gross_wage, tax_brackets):
	"""Calculates tax payable based on the brackets provided."""
	for _, row in tax_brackets.iterrows():
		min_income = row['Remuneration 1']
		max_income = row['Remuneration 2']
		if min_income <= gross_wage <= max_income:
			return row['Under 65']
	return 0

# --- TAX CACULATION ---

def tax(PAYROLL_FILE):
	# 1. Recalculate initial formulas in Payroll file
	recalculate_excel_formulas(PAYROLL_FILE)

	# 2. Read in data
	df_tax = pd.read_excel(TAX_RATES_FILE)
	df_payroll = pd.read_excel(PAYROLL_FILE)

	# Clean tax columns
	for col in['Remuneration 1', 'Remuneration 2', 'Under 65']:
		if col in df_tax.columns:
			df_tax[col] = df_tax[col].apply(clean_currency)

	# 3. Collect employee names and gross wages (Assumes specific row/col structure)
	# Using df_payroll.iloc/loc.
	# skipping first 2 and last 1 columns
	employees = df_payroll.columns[2:-1]
	gross_wages = df_payroll.loc[20].values[2:-1]

	# Combine duplicates (e.g., "John 1" gets added to "John")
	employee_totals = {}

	for emp, wage in zip(employees, gross_wages):		
		# Strip trailing " 1" (or " 2") using regex to find base name
		base_name = re.sub(r'\s*\.?\d+$', '', str(emp)).strip()
		
		if base_name in employee_totals:
			employee_totals[base_name] += wage
		else:
			employee_totals[base_name] = wage
	
	# Clean currency to int
	for name, total_wage in employee_totals.items():
		employee_totals[name] = clean_currency(total_wage)

	# 4. Calculate tax for each person
	results = {}
	for name, gross_wage in employee_totals.items():
		tax_payable = get_tax_amount(gross_wage, df_tax)
		results[name] = {'Gross Wage': gross_wage, 'Tax Payable': tax_payable}

	# 5. Create and save Excel sheet with results
	# Pandas handles the transposition and header writing automatically
	results_df = pd.DataFrame.from_dict(results, orient='index')
	results_df.index.name = 'Employee Name'
	results_df = results_df.T  # Transpose to match old layout
	
	if os.path.isfile(TAX_RESULTS):
		os.remove(TAX_RESULTS)
		
	results_df.to_excel(TAX_RESULTS, sheet_name='Results')

	# 6. Update Payroll File
	wb = load_workbook(PAYROLL_FILE)
	ws = wb.active

	# data_only=True evaluates the formulas so we can read the raw UIF values
	wb_dot = load_workbook(PAYROLL_FILE, data_only=True)
	ws_dot = wb_dot.active

	names_done = set()

	for col in range(3, ws_dot.max_column):
		col_letter = get_column_letter(col)
		name = ws_dot[f'{col_letter}1'].value
		uif = ws_dot[f'{col_letter}23'].value
		
		tax_amt = 0

		if uif is not None and uif > 0:
			# Look up the base name in case the column header is "John 1"
			base_name = re.sub(r'\s*\.?\d+$', '', str(name)).strip()

			if base_name not in names_done:
				tax_amt = results.get(base_name, {}).get('Tax Payable', 0)
				names_done.add(base_name)

		ws[f'{col_letter}30'] = tax_amt

	wb.save(PAYROLL_FILE)
	wb.close()
	wb_dot.close()

	# 7. Final recalculation of formulas
	recalculate_excel_formulas(PAYROLL_FILE)