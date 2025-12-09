import sqlite3
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from tkinter import messagebox

# FUNCTIONS TO INTERACT WITH EMPLOYEE INFOMATION
def add_employees(ename, fname, sname, id):
	try:
		# Check to see if non english name
		if fname == '':
			fname = '0'

		con = sqlite3.connect("wageTimes.db")
		c = con.cursor()

		query = """INSERT INTO employeeNames (englishName, fullName, Surname, idPass)
				VALUES (?, ?, ?, ?)"""
		
		c.execute(query, (ename, fname, sname, id))

		con.commit()
		con.close()
	except Exception as error:
		messagebox.showerror('Error Add Employee', error)

def search_employees(search):
	try:
		con = sqlite3.connect("wageTimes.db")
		c = con.cursor()

		c.execute(f"""SELECT englishName,
						fullName,
						Surname,
						idPass
					FROM
						employeeNames
					WHERE
						englishName LIKE '%{search}%'""")
		
		records = c.fetchall()

		con.commit()
		con.close()

		return records
	except Exception as error:
		messagebox.showerror('Error Search Employee', error)

def update_employees(ename, fname, sname, id):
	try:
		# Check to see if non english name
		if fname == '':
			fname = '0'

		con = sqlite3.connect("wageTimes.db")
		c = con.cursor()

		c.execute(f'''UPDATE employeeNames SET
						englishName = :ename,
						fullName = :fname,
						surname = :sname

						WHERE idPass = :id''',
						{
							'ename' : ename,
							'fname' : fname,
							'sname' : sname,
							'id' : id
						})

		con.commit()
		con.close()
	except Exception as error:
		messagebox.showerror('Error Update Employee', error)

def delete_employees(id):
	try:
		con = sqlite3.connect("wageTimes.db")
		c = con.cursor()

		c.execute(f'''DELETE FROM employeeNames WHERE idPass = :id''',
						{
							'id' : id
						})

		con.commit()
		con.close()
	except Exception as error:
		messagebox.showerror('Error Delete Employee', error)


# add bulk import
# disable id on update

# GENERATE PAY SLIPS
def gen_payslips():
	# Get employee information Full Name and ID/Passport
	employee_names_file = 'Templates/Employee_Names.csv'
	employee_info = pd.read_csv(employee_names_file)
	employee_list = employee_info.values.tolist()

	employee_name_info = {}
	for x in employee_list:
		emp_eng_name = str(x[0]).strip()
		emp_name = str(x[1]).strip()
		emp_surname = str(x[2]).strip()
		emp_id = str(x[3]).strip()

		employee_name_info[emp_eng_name] = [emp_eng_name, emp_name, emp_surname, emp_id]

	# Read in wxcel workbook
	df = pd.read_excel('Payroll/payroll.xlsx')

	# Get column names
	col_names = df.columns.tolist()

	# Get date of wages
	date = col_names[0]

	# Get info of first column headings 
	headings = df[df.columns[0]].tolist()
	headings[:0] = ['name']

	# Get names of employees
	names = col_names[2:-1]  # Filter list remove first 2 elements and last element

	employees = []
	for e in names:
		if 'Null' not in e:
			employees.append(e)

	# Get employee info and loop through and create payslips
	for name in employees:
		index = col_names.index(name)
		emp_info = df.get(col_names[index]).tolist()
		emp_info[:0] = [name]

		# # Zip info into dict
		pay_info = dict(zip(headings, emp_info))

		# Create Workbooks with payslips
		wb = load_workbook('Templates/Payslip_Template.xlsx')
		ws = wb.active

		# Name of employee and ID/Passport
		ws['B6'] = 'Name'
		ws['B7'] = 'ID/Passport'

		if name[-2:] == '.1':
			if employee_name_info[name[:-2].strip()][1] == '0':
				ws['C6'] = f'{employee_name_info[name[:-2].strip()][0]} {employee_name_info[name[:-2].strip()][2]}'
				ws['C7'] = employee_name_info[name[:-2].strip()][3]
			else:
				ws['C6'] = f'{employee_name_info[name[:-2].strip()][1]} {employee_name_info[name[:-2].strip()][2]}'
				ws['C7'] = employee_name_info[name[:-2].strip()][3]

		elif employee_name_info[name.strip()][1] == '0':
			ws['C6'] = f'{employee_name_info[name.strip()][0]} {employee_name_info[name.strip()][2]}'
			ws['C7'] = employee_name_info[name.strip()][3]
		else:
			ws['C6'] = f'{employee_name_info[name.strip()][1]} {employee_name_info[name.strip()][2]}'
			ws['C7'] = employee_name_info[name.strip()][3]

		# Occupation
		ws['B8'] = 'Occupation'
		ws['C8'] = pay_info['Occupation']

		# Fortnight ending
		ws['B9'] = 'Fortnight Ending'
		ws['C9'] = date

		# Hour, Rate, Amount
		ws['C11'] = 'Hours'
		ws['D11'] = 'Rate'
		ws['E11'] = 'Amount'

		# Ordinary Time
		ws['B12'] = 'Ordinary Time'
		ws['C12'] = pay_info['HRS']
		ws['D12'] = pay_info['N_RATE']
		ws['E12'] = '=C12*D12'

		# Overtime
		ws['B13'] = 'Overtime'
		ws['C13'] = pay_info['O/T HRS']
		ws['D13'] = pay_info['OT_RATE']
		ws['E13'] = '=C13*D13'

		# Sunday Times
		ws['B14'] = 'Sunday Times'
		ws['C14'] = pay_info['SUNDAY']
		ws['D14'] = pay_info['S_RATE']
		ws['E14'] = '=C14*D14'

		# Public Holiday
		ws['B15'] = 'Public Holidays'
		ws['C15'] = pay_info['PUB HOL HRS']
		ws['D15'] = pay_info['PH_RATE']
		ws['E15'] = '=C15*D15'

		# Bonus
		ws['B16'] = 'Bonus'
		ws['E16'] = pay_info['BONUS']

		# Sick Leave / Leave
		ws['B17'] = 'Sick Leave / Leave'
		ws['E17'] = pay_info['SICK / LEAVE']

		# Medical Allowance
		ws['B18'] = 'Medical'
		ws['E18'] = pay_info['MEDICAL']

		# Standard Hours
		ws['B19'] = 'Standard Hours'
		ws['E19'] = pay_info['STANDARD HRS']

		# Total Wage
		ws['B20'] = 'Total Wage'
		ws['E20'] = pay_info['TOTAL WAGE']

		# Deduction Heading
		ws['B22'] = 'Deductions'

		# UIF
		ws['B23'] = 'UIF'
		ws['E23'] = pay_info['UIF']

		# Mibco
		ws['B24'] = 'Mibco'
		ws['E24'] = pay_info['MIBCO']

		# Uniforms
		ws['B25'] = 'Uniforms'
		ws['E25'] = pay_info['UNIFORMS']

		# Union
		ws['B26'] = 'Union'
		ws['E26'] = pay_info['UNION']

		# Advance
		ws['B27'] = 'Advance'
		ws['E27'] = pay_info['ADVANCES']

		# Prov Fund
		ws['B28'] = 'Prov Fund'
		ws['E28'] = pay_info['PROV FUND']

		# PAYE
		ws['B29'] = 'PAYE'
		ws['E29'] = pay_info['PAYE']

		# PAYE REPAY
		ws['B30'] = 'PAYE REPAY'
		ws['E30'] = pay_info['PAYE REPAYME']

		# Shortages
		ws['B31'] = 'Shortages'
		ws['E31'] = pay_info['SHORTAGES']

		# Net Wage
		ws['B33'] = 'NET WAGE'
		ws['E33'] = pay_info['NET WAGE']

		# FORMATING
		# Border         
		thin = Side(border_style="thin")
		thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
		
		# Font for headings    
		font_headers = Font(name='Arial', size=10, bold=True)
		
		# Column Dimentions and formating (Title, name, occupation and week ending)
		ws.column_dimensions['B'].width = 17.04
		# ws['C6'].font = font_headers
		
		# Merge cells (Name, ID, occupation, week ending)
		ws.merge_cells('C6:E6')
		ws["C6"].alignment = Alignment(horizontal="center")
		ws['C6'].font = font_headers
		
		ws.merge_cells('C7:E7')
		ws["C7"].alignment = Alignment(horizontal="center")
		
		ws.merge_cells('C8:E8')
		ws["C8"].alignment = Alignment(horizontal="center")
		
		ws.merge_cells('C9:E9')
		ws["C9"].alignment = Alignment(horizontal="center")
		
		# ws.merge_cells('C10:E10')
		# ws["C10"].alignment = Alignment(horizontal="center")
		
		# Bold Headings
		ws['B20'].font = font_headers
		ws['B22'].font = font_headers
		ws['B33'].font = font_headers

		# Format whole payslip
		for row in range(6,34):
			for col in range(2,6):
				ws.cell(row,col).border = thin_border
		
		ws.column_dimensions['E'].width = 9.65
		
		for row in range(12,34):
			for col in range(4,6):
				ws.cell(row,col).number_format = 'R #,##0.00'

		if name[-2:] == '.1':
			wb.save(f'Payslips/{name[:-2].strip()} Bakery.xlsx')
		else:
			wb.save(f'Payslips/{name}.xlsx')