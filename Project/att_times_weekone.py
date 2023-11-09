# ###############################################################################################
# ATTENDENTS WEEK 1 - CALCULATE ROSTER VS CLOCK TIME IN EXCEL
# ###############################################################################################
import os
from datetime import datetime, time
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import sqlite3


# ADD CALULATIONS OF HOURS FOR WEEK AND SUNDAY HOURS

# ==============================================================================
# Calculate Clock in and Clock out times
# ==============================================================================

def att_times_weekone():
    wb = load_workbook("Wage Times.xlsx")
    ws = wb.active

    i = 0
    for x in range(ws.max_row + 1):
        tti = 0
        tto = 0

        name = ws.cell(row=2 + i, column=1).value

        day = ws.cell(row=2 + i, column=3).value

        time_in = ws.cell(row=2 + i, column=5).value
        clock_in = ws.cell(row=2 + i, column=7).value

        time_out = ws.cell(row=2 + i, column=6).value
        clock_out = ws.cell(row=2 + i, column=8).value

        if name == None:
            i += 1

        elif time_in > 0 and clock_in == None or time_out > 0 and clock_out == None:
            hours = 'No Clock'
            ws.cell(row=2 + i, column=11, value=hours)

            i += 1

        elif day == "Sunday":
            if time_in == 18:
                ti = time(time_in).strftime("%H:%M")
                ci = clock_in

                if ci > ti:
                    if ci[-2:] < "05":
                        tti = 24.0 - float(ci[0:2])
                    elif ci[-2:] < "15":
                        tti = 24.0 - (float(ci[0:2]) + 0.25)
                    elif ci[-2:] < "30":
                        tti = 24.0 - (float(ci[0:2]) + 0.30)
                    elif ci[-2:] < "45":
                        tti = 24.0 - (float(ci[0:2]) + 0.75)
                    elif ci[-2:] > "45":
                        tti = 24.0 - (float(ci[0:2]) + 1.0)
                else:
                    tti = 24.0 - float(time_in)

                hours = tti
                ws.cell(row=2 + i, column=10, value=hours)

                i += 1
            
            elif clock_in == None and clock_out == None:
                hours = 0
                ws.cell(row=2 + i, column=10, value=hours)

                i += 1

            elif clock_in == None:
                # Calculate time out
                to = time(time_out).strftime("%H:%M")
                co = clock_out

                if co < to:
                    if co[-2:] < "05":
                        tto = float(co[0:2]) + 0.00
                    elif co[-2:] < "15":
                        tto = float(co[0:2]) + 0.25
                    elif co[-2:] < "30":
                        tto = float(co[0:2]) + 0.30
                    elif co[-2:] < "45":
                        tto = float(co[0:2]) + 0.75
                    elif co[-2:] > "45":
                        tto = (float(co[0:2]) + 1) + 0.00
                else:
                    tto = int(time_out)

                hours = float(tto)
                ws.cell(row=2 + i, column=10, value=hours)

                i += 1

            else:
                ti = time(time_in).strftime("%H:%M")
                ci = clock_in

                to = time(time_out).strftime("%H:%M")
                co = clock_out

                # Calculate time in
                if ci > ti:
                    if ci[-2:] < "05":
                        tti = float(ci[0:2])
                    elif ci[-2:] < "15":
                        tti = float(ci[0:2]) + 0.25
                    elif ci[-2:] < "30":
                        tti = float(ci[0:2] + 0.30)
                    elif ci[-2:] < "45":
                        tti = float(ci[0:2]) + 0.75
                    elif ci[-2:] > "45":
                        tti = (float(ci[0:2]) + 1) + 0.00
                else:
                    tti = float(time_in)

                # Calculate time out
                if co < to:
                    if co[-2:] < "05":
                        tto = float(co[0:2]) + 0.00
                    elif co[-2:] < "15":
                        tto = float(co[0:2]) + 0.25
                    elif co[-2:] < "30":
                        tto = float(co[0:2]) + 0.30
                    elif co[-2:] < "45":
                        tto = float(co[0:2]) + 0.75
                    elif co[-2:] > "45":
                        tto = (float(co[0:2]) + 1) + 0.00
                else:
                    tto = int(time_out)

                # calculate hours worked
                hours = float(tto) - float(tti)
                ws.cell(row=2 + i, column=10, value=hours)

            i += 1

        elif clock_in == None and clock_out == None:
            hours = 0
            ws.cell(row=2 + i, column=9, value=hours)

            i += 1

        elif time_in == 18:
            ti = time(time_in).strftime("%H:%M")
            ci = clock_in

            if ci > ti:
                if ci[-2:] < "05":
                    tti = 24.0 - float(ci[0:2])
                elif ci[-2:] < "15":
                    tti = 24.0 - (float(ci[0:2]) + 0.25)
                elif ci[-2:] < "30":
                    tti = 24.0 - (float(ci[0:2]) + 0.30)
                elif ci[-2:] < "45":
                    tti = 24.0 - (float(ci[0:2]) + 0.75)
                elif ci[-2:] > "45":
                    tti = 24.0 - (float(ci[0:2]) + 1.0)
            else:
                tti = 24.0 - float(time_in)

            hours = tti
            ws.cell(row=2 + i, column=9, value=hours)

            i += 1

        elif clock_in == None:
            # Calculate time out
            to = time(time_out).strftime("%H:%M")
            co = clock_out

            if co < to:
                if co[-2:] < "05":
                    tto = float(co[0:2]) + 0.00
                elif co[-2:] < "15":
                    tto = float(co[0:2]) + 0.25
                elif co[-2:] < "30":
                    tto = float(co[0:2]) + 0.30
                elif co[-2:] < "45":
                    tto = float(co[0:2]) + 0.75
                elif co[-2:] > "45":
                    tto = (float(co[0:2]) + 1) + 0.00
            else:
                tto = int(time_out)

            hours = float(tto)
            ws.cell(row=2 + i, column=9, value=hours)

            i += 1

        else:
            ti = time(time_in).strftime("%H:%M")
            ci = clock_in

            to = time(time_out).strftime("%H:%M")
            co = clock_out

            # Calculate time in
            if ci > ti:
                if ci[-2:] < "05":
                    tti = float(ci[0:2])
                elif ci[-2:] < "15":
                    tti = float(ci[0:2]) + 0.25
                elif ci[-2:] < "30":
                    tti = float(ci[0:2] + 0.30)
                elif ci[-2:] < "45":
                    tti = float(ci[0:2]) + 0.75
                elif ci[-2:] > "45":
                    tti = (float(ci[0:2]) + 1) + 0.00
            else:
                tti = float(time_in)

            # Calculate time out
            if co < to:
                if co[-2:] < "05":
                    tto = float(co[0:2]) + 0.00
                elif co[-2:] < "15":
                    tto = float(co[0:2]) + 0.25
                elif co[-2:] < "30":
                    tto = float(co[0:2]) + 0.30
                elif co[-2:] < "45":
                    tto = float(co[0:2]) + 0.75
                elif co[-2:] > "45":
                    tto = (float(co[0:2]) + 1) + 0.00
            else:
                tto = int(time_out)

            # calculate hours worked
            hours = float(tto) - float(tti)
            ws.cell(row=2 + i, column=9, value=hours)

            i += 1

    wb.save("Wage Times.xlsx")
    wb.close()







# Calculate total hours for week add to excel
wb = load_workbook("Wage Times.xlsx")
ws = wb.active

i = 0
total = 0
total_s = 0

# print(ws.max_row)
for x in range(ws.max_row):
    name = ws.cell(row=2 + i, column=1).value
    n = ws.cell(row=2 + i - 1, column=1).value
    day = ws.cell(row=2 + i, column=3).value
    hours = ws.cell(row=2 + i, column=9).value
    hours_s = ws.cell(row=2 + i, column=10).value
    
    # Check if name is true
    if name:
        if day == 'Sunday':
            total_s += hours_s
        elif hours == None:
            total += 0
        else:
            total += hours
        
        i += 1
    
    elif "Total" in n:
        i += 1

    else:
        ws.cell(row=2 + i, column=1, value= n + " " + "Total")
        ws.cell(row=2 + i, column=9, value=total)
        ws.cell(row=2 + i, column=10, value=total_s)
        
        total = 0            
        total_s = 0  

        i += 1

wb.save("Wage Times.xlsx")
wb.close()