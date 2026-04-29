import shutil
import pandas as pd
import database as db
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from config import (DYNAMIC_FILE_LOC, PAYSLIP_TEMPLATE, PAYSLIP_FOLDER, COPY_FOLDER, WAGE_TIMES_FILE, 
                    ATT_ROSTER_FILE, CAS_ROSTER_FILE, CARWASH_FILE, TAX_RESULTS, CARWASH_HOURS_FILE)


# =================================================================================================
# GENERATE PAYSLIPS
# =================================================================================================

# --- HELPER FUNCTIONS ---
def get_employee_details(name_key, emp_info_dict):
	"""Helper function to parse and format the employee's name and ID."""
	info = emp_info_dict.get(name_key.strip())
	if not info:
		return "Unknown", "Unknown"
	
	# info[1] seems to toggle whether to use index 0 or 1 for the first name
	first_name = info[0] if info[1] == '0' else info[1]
	last_name = info[2]
	id_passport = info[3]
	
	return f"{first_name} {last_name}", id_passport

# --- MAIN FUNCTION ---
def gen_payslips():
	# Get payroll file
	PAYROLL_FILE = DYNAMIC_FILE_LOC('Payroll')

	# Get employee information
	employee_name_info = db.get_emp_info()

	# Read in payroll excel workbook
	df = pd.read_excel(PAYROLL_FILE)
	col_names = df.columns.tolist()
	date = col_names[0]
	headings = df.iloc[:, 0].tolist() # Get the first column data

	# Filter columns to valid employee names only
	employees = [e for e in col_names[2:-1] if 'Null' not in str(e)]

	# Pre-define Excel styles
	thin_border = Border(top=Side(border_style="thin"), left=Side(border_style="thin"),
						 right=Side(border_style="thin"), bottom=Side(border_style="thin"))
	font_headers = Font(name='Arial', size=10, bold=True)
	center_align = Alignment(horizontal="center")

	for name in employees:
		pay_info = dict(zip(headings, df[name].tolist()))
		
		# Determine if it is cashier/baker's bakery times and clean the name string
		is_bakery = name.endswith('.1')
		clean_name = name[:-2].strip() if is_bakery else name.strip()

		# Get Full Name and ID
		full_name, id_passport = get_employee_details(clean_name, employee_name_info)

		# Load Template
		wb = load_workbook(PAYSLIP_TEMPLATE)
		ws = wb.active

		# 1. Populate standard text, data, and formulas using mapping 
		cell_mappings = {
			# Dynamic Info
			'C6': full_name, 'C7': id_passport, 'C8': pay_info.get('Occupation', ''), 'C9': date,
			'C12': pay_info.get('HRS'),         'D12': pay_info.get('N_RATE'),
			'C13': pay_info.get('O/T HRS'),     'D13': pay_info.get('OT_RATE'),
			'C14': pay_info.get('SUNDAY'),      'D14': pay_info.get('S_RATE'),
			'C15': pay_info.get('PUB HOL HRS'), 'D15': pay_info.get('PH_RATE'),
			'E16': pay_info.get('BONUS'),       'E17': pay_info.get('SICK / LEAVE'),
			'E18': pay_info.get('MEDICAL ALLOW'),'E19': pay_info.get('STANDARD HRS'),
			'E20': pay_info.get('TOTAL WAGE'),  'E23': pay_info.get('UIF'),
			'E24': pay_info.get('MIBCO'),       'E25': pay_info.get('UNIFORMS'),
			'E26': pay_info.get('UNION'),       'E27': pay_info.get('ADVANCES'),
			'E28': pay_info.get('PROV FUND'),   'E29': pay_info.get('PAYE'),
			'E30': pay_info.get('MEDICAL AID'), 'E31': pay_info.get('SHORTAGES'),
			'E33': pay_info.get('NET WAGE'),

			# Formulas
			'E12': '=C12*D12', 'E13': '=C13*D13', 'E14': '=C14*D14', 'E15': '=C15*D15',
		}

		# Apply mapped values
		for cell, value in cell_mappings.items():
			ws[cell] = value

		"""
		1 .Updated Payslip function to use template and not recreate excel every loop.
		2. Any change now made on template and update cell mapping.		
		"""
		# # 2. Apply Formatting
		# ws.column_dimensions['B'].width = 17.04
		# ws.column_dimensions['E'].width = 9.65

		# # Merge and align header cells
		# for row in range(6, 10): # Covers rows 6, 7, 8, 9
		# 	ws.merge_cells(f'C{row}:E{row}')
		# 	ws[f'C{row}'].alignment = center_align
		
		# # Apply specific fonts
		# ws['C6'].font = font_headers
		# for cell in['B20', 'B22', 'B33']:
		# 	ws[cell].font = font_headers

		# # Apply borders (Rows 6 to 33, Columns B to E)
		# for row in range(6, 34):
		# 	for col in range(2, 6):
		# 		ws.cell(row, col).border = thin_border
		
		# # Apply currency format (Rows 12 to 33, Columns D & E)
		# for row in range(12, 34):
		# 	for col in range(4, 6):
		# 		ws.cell(row, col).number_format = 'R #,##0.00'

		# 3. Save Workbook
		suffix = ' Bakery' if is_bakery else ''
		wb.save(f'{PAYSLIP_FOLDER}/{clean_name}{suffix}.xlsx')

# =================================================================================================
# BACKUP FILES
# =================================================================================================

def copy_files():
    # Define the destination
    dest_folder = COPY_FOLDER
    
	# Get payroll file
    PAYROLL_FILE = DYNAMIC_FILE_LOC('Payroll')
    
    # List of source files
    files_to_copy = [
        WAGE_TIMES_FILE,
        PAYROLL_FILE,
        ATT_ROSTER_FILE,
        CAS_ROSTER_FILE,
        CARWASH_FILE,
        TAX_RESULTS,
        CARWASH_HOURS_FILE
    ]
    
    for file_path in files_to_copy:
        try:
            shutil.copy2(file_path, dest_folder)
        except FileNotFoundError:
            raise FileExistsError("Could Not Find File")
        except Exception:
            raise