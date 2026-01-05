from openpyxl import load_workbook
import sqlite3

def carwash_times():
	# Load workbook carwash times
	wb = load_workbook("Carwash Times/Carwash Times.xlsx", data_only=True)
	ws = wb['Times']

	# Gather info from rows
	rows = ws.iter_rows(min_row=3, max_row=10, min_col=12, max_col=16)	
	extra_time = ws.iter_rows(min_row=14, max_row=21, min_col=12, max_col=16)

	# Put all info in dict and list
	carwash_times = {}
	carwash_extra_time = []

	for row in rows:
		if row[0].value != None:
			x = [str(row[0].value), str(row[1].value), str(row[2].value), str(row[3].value)]
			carwash_times[str(row[1].value)] = x

	for x in extra_time:
		x = [str(x[0].value), str(x[1].value), str(x[2].value), str(x[3].value)]
		carwash_extra_time.append(x)

	for x in carwash_times:
		for y in carwash_extra_time:
			if x == y[1]:
				carwash_times[x].append(y[3])

	# CREATE DATABASE AND ADD TIMES
	# Connect to database
	con = sqlite3.connect("wageTimes.db")
	c = con.cursor()
	# Create table
	c.execute(
		"""CREATE TABLE IF NOT EXISTS carwashTotal (
			name TEXT,
			badge TEXT,
			normal TEXT,
			sunday TEXT,
			public TEXT,
			extra TEXT
			)"""
	)

	# Add week one data to table
	query = """INSERT INTO carwashTotal (
			name,
			badge,
			normal,
			sunday,
			public,
			extra
			)
			VALUES (?, ?, ?, ?, ?, ?)"""

	# Add times to database
	for x in carwash_times:
		c.execute(query, (carwash_times[x][0], carwash_times[x][1], carwash_times[x][2], carwash_times[x][3], '0', carwash_times[x][4]))

	con.commit()
	con.close()