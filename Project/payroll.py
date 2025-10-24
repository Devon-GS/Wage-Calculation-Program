from openpyxl import Workbook
from openpyxl import load_workbook
import carwash_db as car
import sqlite3

def payroll():
    # Run Carwash to database
    car.carwash_times()

    # GET ALL HOURS FROM DATA BASE
    con = sqlite3.connect("wageTimes.db")
    c = con.cursor()

    # Attendent Times
    c.execute("SELECT name, badge, normal, sunday, public FROM attTotal")
    records_att = c.fetchall()

    # Cashier Times
    c.execute("SELECT name, badge, normal, sunday, public, cashier FROM cashierTotal")
    records_cash = c.fetchall()

    # Carwash times
    c.execute("SELECT name, badge, normal, sunday, public, extra FROM carwashTotal")
    records_car = c.fetchall()

    con.commit()
    con.close()

    # JOIN ALL INFO FROM DATABASE INTO ONE TABLE
    total_records = []

    for rec in records_att:
        # convert rec to list
        convert = list(rec)
        convert.append(0)
        total_records.append(convert)

    for rec in records_cash:
        total_records.append(rec)

    for rec in records_car:
        total_records.append(rec)

    # COPY HOURS FROM TOTAL INFO TO PAYROLL.XLSX
    wb = load_workbook("Payroll/Payroll.xlsx")
    ws = wb['Wages']

    # Iterate over payrol columns
    columns = ws.iter_cols(min_row=1, min_col=3)

    # If badge number in payroll match info copy hours
    # col[2] = normal pay
    # col[11] = sunday pay
    # col[14] = public pay
    for col in columns:
        for r in total_records:
            if 'c' in str(col[1].value) and str(col[1].value)[:-1] == str(r[1]):
                col[2].value = float(r[5])
                col[11].value = 0.00
                col[14].value = 0.00
            elif 'b' in str(col[1].value) and str(col[1].value)[:-1] == str(r[1]):
                col[2].value = float(r[2])
                col[11].value = float(r[3])
                col[14].value = float(r[4])
            elif col[1].value == int(r[1]):
                    col[2].value = float(r[2])
                    col[11].value = float(r[3])
                    col[14].value = float(r[4])
                    if float(r[5]) != None and float(r[5]) != 0.0 and float(r[5]) != 1.0:
                         col[20].value = float(r[5])

    wb.save("Payroll/Payroll.xlsx")
    wb.close()